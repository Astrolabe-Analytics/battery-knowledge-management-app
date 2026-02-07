#!/usr/bin/env python3
"""
Streamlit web interface: Paper Library + RAG Query System
Pure UI layer - all business logic delegated to lib.rag module

Now uses improved retrieval pipeline with:
- Query expansion (Claude expands queries with related terms)
- Hybrid search (combines vector similarity + BM25 keyword search)
- Reranking (retrieves 15 candidates, reorders by relevance, returns top 5)
"""

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

import os
import sys
import json
import re
import html
import time
import requests
import streamlit as st
import pandas as pd
from datetime import datetime
from typing import Dict, Any
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode
from pathlib import Path

# Fix Windows console encoding for Unicode characters
if sys.platform == 'win32':
    import codecs
    if hasattr(sys.stdout, 'buffer'):
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    if hasattr(sys.stderr, 'buffer'):
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Import backend
from lib import rag, read_status, query_history, theme, styles, collections, gap_analysis, semantic_scholar, cached_operations


def clean_html_from_text(text: str) -> str:
    """
    Remove HTML entities and tags from text.
    Converts &lt;sub&gt;4&lt;/sub&gt; to just "4"
    """
    if not text:
        return text

    # First, unescape HTML entities (&lt; -> <, &gt; -> >, etc.)
    text = html.unescape(text)

    # Then strip all HTML tags
    text = re.sub(r'<[^>]+>', '', text)

    return text


# Page config
st.set_page_config(
    page_title="Astrolabe Research Library",
    page_icon="⚛️",
    layout="wide",
    initial_sidebar_state="expanded"
)


def load_theme_preference():
    """Load theme preference from settings file."""
    settings_file = Path("data/settings.json")
    if settings_file.exists():
        try:
            with open(settings_file, 'r') as f:
                settings = json.load(f)
                return settings.get('theme', 'light')
        except:
            return 'light'
    return 'light'


def load_settings() -> dict:
    """Load settings from settings file."""
    settings_file = Path("data/settings.json")
    if settings_file.exists():
        try:
            with open(settings_file, 'r') as f:
                return json.load(f)
        except:
            pass
    return {}


def save_settings(settings: dict):
    """Save settings to settings file."""
    settings_file = Path("data/settings.json")
    settings_file.parent.mkdir(parents=True, exist_ok=True)
    with open(settings_file, 'w') as f:
        json.dump(settings, f, indent=2)


def save_theme_preference(theme: str):
    """Save theme preference to settings file."""
    settings = load_settings()
    settings['theme'] = theme
    save_settings(settings)


def get_api_key():
    """Get Anthropic API key from environment or user input (UI layer)."""
    api_key = rag.get_api_key_from_env()
    if not api_key:
        api_key = st.session_state.get("anthropic_api_key")
    if not api_key:
        with st.sidebar:
            st.warning("⚠️ Anthropic API key required for queries")
            api_key = st.text_input(
                "Enter your Anthropic API key:",
                type="password",
                help="Get your API key from https://console.anthropic.com/"
            )
            if api_key:
                st.session_state.anthropic_api_key = api_key
                st.rerun()
    return api_key


def query_crossref_for_metadata(doi: str) -> dict:
    """Query CrossRef API for canonical metadata using DOI."""
    try:
        url = f"https://api.crossref.org/works/{doi}"
        headers = {
            'User-Agent': 'BatteryPaperLibrary/1.0 (mailto:researcher@example.com)'
        }
        response = requests.get(url, headers=headers, timeout=10)

        if response.status_code == 200:
            data = response.json()
            message = data.get('message', {})

            metadata = {}

            # Title
            titles = message.get('title', [])
            if titles:
                metadata['title'] = titles[0]

            # Authors (format as "Last, First")
            authors = []
            for author in message.get('author', []):
                given = author.get('given', '')
                family = author.get('family', '')
                if family:
                    if given:
                        authors.append(f"{family}, {given}")
                    else:
                        authors.append(family)
            metadata['authors'] = authors[:10]  # Limit to 10

            # Year
            published = message.get('published-print') or message.get('published-online')
            if published and 'date-parts' in published:
                date_parts = published['date-parts'][0]
                if date_parts:
                    metadata['year'] = str(date_parts[0])

            # Journal
            container_titles = message.get('container-title', [])
            if container_titles:
                metadata['journal'] = container_titles[0]

            # Abstract
            metadata['abstract'] = message.get('abstract', '')

            # Author keywords
            metadata['author_keywords'] = message.get('keywords', [])

            # Volume, Issue, Pages
            metadata['volume'] = message.get('volume', '')
            metadata['issue'] = message.get('issue', '')
            metadata['pages'] = message.get('page', '')

            # References
            metadata['references'] = message.get('reference', [])

            return metadata
        else:
            return {}
    except Exception as e:
        st.error(f"CrossRef API error: {e}")
        return {}


def save_metadata_only_paper(doi: str, crossref_metadata: dict) -> str:
    """Save metadata-only paper to ChromaDB and metadata.json"""
    import chromadb

    safe_doi = doi.replace('/', '_').replace('.', '_')
    filename = f"doi_{safe_doi}.pdf"

    # Save to metadata.json
    metadata_file = Path("data/metadata.json")
    metadata_file.parent.mkdir(parents=True, exist_ok=True)

    all_metadata = {}
    if metadata_file.exists():
        with open(metadata_file, 'r', encoding='utf-8') as f:
            all_metadata = json.load(f)

    all_metadata[filename] = {
        'filename': filename,
        'title': crossref_metadata.get('title', 'Unknown Title'),
        'authors': crossref_metadata.get('authors', []),
        'year': crossref_metadata.get('year', ''),
        'journal': crossref_metadata.get('journal', ''),
        'doi': doi,
        'chemistries': [],
        'topics': [],
        'application': 'general',
        'paper_type': 'experimental',
        'metadata_only': True,
        'date_added': datetime.now().isoformat(),
        'abstract': crossref_metadata.get('abstract', ''),
        'author_keywords': crossref_metadata.get('author_keywords', []),
        'volume': crossref_metadata.get('volume', ''),
        'issue': crossref_metadata.get('issue', ''),
        'pages': crossref_metadata.get('pages', ''),
        'source_url': '',
        'notes': '',
        'references': crossref_metadata.get('references', [])
    }

    with open(metadata_file, 'w', encoding='utf-8') as f:
        json.dump(all_metadata, f, indent=2, ensure_ascii=False)

    # Add to ChromaDB using the DatabaseClient to ensure consistency
    from lib.rag import DatabaseClient

    # First clear any cached collection to force a fresh connection
    DatabaseClient.clear_cache()

    # Now get a fresh collection reference
    collection = DatabaseClient.get_collection()

    doc_id = f"{filename}_metadata_only"
    try:
        collection.delete(ids=[doc_id])
    except:
        pass

    collection.add(
        documents=[f"Metadata-only: {crossref_metadata.get('title', '')}. DOI: {doi}"],
        metadatas=[{
            'filename': filename,
            'page_num': 0,
            'section_name': 'metadata_only',
            'title': crossref_metadata.get('title', ''),
            'authors': ';'.join(crossref_metadata.get('authors', [])) if crossref_metadata.get('authors') else '',
            'year': crossref_metadata.get('year', ''),
            'journal': crossref_metadata.get('journal', ''),
            'doi': doi,
            'chemistries': '',
            'topics': '',
            'application': 'general',
            'paper_type': 'experimental',
            'abstract': crossref_metadata.get('abstract', ''),
            'author_keywords': ';'.join(crossref_metadata.get('author_keywords', [])),
            'volume': crossref_metadata.get('volume', ''),
            'issue': crossref_metadata.get('issue', ''),
            'pages': crossref_metadata.get('pages', ''),
            'date_added': datetime.now().isoformat(),
            'source_url': ''
        }],
        ids=[doc_id]
    )

    # Clear cache again so next get_paper_library() call sees the new paper
    DatabaseClient.clear_cache()

    return filename


def add_paper_with_pdf_search(doi: str, title: str, authors: str, year: str, url: str = '') -> dict:
    """
    Add paper to library, attempting to find and download open access PDF.

    Args:
        doi: Paper DOI
        title: Paper title
        authors: Paper authors string
        year: Publication year
        url: Paper URL (optional)

    Returns:
        Dict with 'success', 'message', 'pdf_found' keys
    """
    try:
        # Step 1: Get metadata from CrossRef if we have DOI
        if doi:
            crossref_metadata = query_crossref_for_metadata(doi)
            if not crossref_metadata or not crossref_metadata.get('title'):
                # Fallback to provided metadata
                crossref_metadata = {
                    'title': title,
                    'authors': [authors] if authors else [],
                    'year': year
                }
        else:
            # No DOI - use provided metadata
            crossref_metadata = {
                'title': title,
                'authors': [authors] if authors else [],
                'year': year
            }

        # Step 2: Try to find open access PDF via Semantic Scholar
        pdf_url = None
        pdf_found = False

        if doi:
            # Search Semantic Scholar for this DOI
            try:
                search_result = semantic_scholar.search_papers(
                    query=f'doi:{doi}',
                    limit=1
                )

                if search_result['success'] and search_result['data']:
                    paper_data = search_result['data'][0]
                    formatted = semantic_scholar.format_paper_for_display(paper_data)

                    if formatted['is_open_access'] and formatted['pdf_url']:
                        pdf_url = formatted['pdf_url']
                        pdf_found = True
            except:
                pass  # Continue even if Semantic Scholar fails

        # Step 3: Try Unpaywall as fallback (if still no PDF)
        if not pdf_found and doi:
            try:
                import requests
                unpaywall_url = f"https://api.unpaywall.org/v2/{doi}?email=user@example.com"
                response = requests.get(unpaywall_url, timeout=10)

                if response.status_code == 200:
                    unpaywall_data = response.json()
                    if unpaywall_data.get('is_oa') and unpaywall_data.get('best_oa_location'):
                        oa_location = unpaywall_data['best_oa_location']
                        if oa_location.get('url_for_pdf'):
                            pdf_url = oa_location['url_for_pdf']
                            pdf_found = True
            except:
                pass  # Continue even if Unpaywall fails

        # Step 4: Download PDF if found
        filename = None
        if pdf_found and pdf_url:
            # Create safe filename
            safe_title = re.sub(r'[^\w\s-]', '', title[:50])
            safe_title = re.sub(r'[-\s]+', '_', safe_title)
            filename = f"{safe_title}.pdf"
            pdf_path = Path("papers") / filename

            download_result = semantic_scholar.download_pdf(pdf_url, pdf_path)

            if not download_result['success']:
                # Download failed, fall back to metadata-only
                pdf_found = False
                filename = None

        # Step 5: Save metadata
        if not filename:
            # Metadata-only
            if doi:
                safe_doi = doi.replace('/', '_').replace('.', '_')
                filename = f"doi_{safe_doi}.pdf"
            else:
                safe_title = re.sub(r'[^\w\s-]', '', title[:50])
                safe_title = re.sub(r'[-\s]+', '_', safe_title)
                filename = f"{safe_title}.pdf"

        # Save to metadata.json
        metadata_file = Path("data/metadata.json")
        metadata_file.parent.mkdir(parents=True, exist_ok=True)

        all_metadata = {}
        if metadata_file.exists():
            with open(metadata_file, 'r', encoding='utf-8') as f:
                all_metadata = json.load(f)

        # Determine PDF status
        pdf_status = "available" if pdf_found else "needs_pdf"

        all_metadata[filename] = {
            'filename': filename,
            'title': crossref_metadata.get('title', title),
            'authors': crossref_metadata.get('authors', [authors] if authors else []),
            'year': crossref_metadata.get('year', year),
            'journal': crossref_metadata.get('journal', ''),
            'doi': doi,
            'chemistries': [],
            'topics': [],
            'application': 'general',
            'paper_type': 'experimental',
            'metadata_only': not pdf_found,
            'pdf_status': pdf_status,
            'date_added': datetime.now().isoformat(),
            'abstract': crossref_metadata.get('abstract', ''),
            'author_keywords': crossref_metadata.get('author_keywords', []),
            'volume': crossref_metadata.get('volume', ''),
            'issue': crossref_metadata.get('issue', ''),
            'pages': crossref_metadata.get('pages', ''),
            'source_url': url,
            'notes': '',
            'references': crossref_metadata.get('references', [])
        }

        with open(metadata_file, 'w', encoding='utf-8') as f:
            json.dump(all_metadata, f, indent=2, ensure_ascii=False)

        # Add to ChromaDB
        from lib.rag import DatabaseClient
        DatabaseClient.clear_cache()
        collection = DatabaseClient.get_collection()

        doc_id = f"{filename}_metadata"
        try:
            collection.delete(ids=[doc_id])
        except:
            pass

        collection.add(
            documents=[f"Metadata: {crossref_metadata.get('title', title)}. DOI: {doi}"],
            metadatas=[{
                'filename': filename,
                'page_num': 0,
                'paper_type': 'experimental',
                'application': 'general',
                'chemistries': '',
                'topics': '',
                'section_name': 'metadata',
                'abstract': crossref_metadata.get('abstract', ''),
                'author_keywords': ';'.join(crossref_metadata.get('author_keywords', [])),
                'title': crossref_metadata.get('title', title),
                'authors': '; '.join(crossref_metadata.get('authors', [])) if isinstance(crossref_metadata.get('authors'), list) else str(crossref_metadata.get('authors', '')),
                'year': str(crossref_metadata.get('year', year)),
                'journal': crossref_metadata.get('journal', ''),
                'doi': doi
            }],
            ids=[doc_id]
        )

        DatabaseClient.clear_cache()

        return {
            'success': True,
            'message': f"Added: {title[:50]}..." if len(title) > 50 else title,
            'pdf_found': pdf_found
        }

    except Exception as e:
        return {
            'success': False,
            'message': f"Failed to add {title[:30]}...: {str(e)}",
            'pdf_found': False
        }


def extract_doi_from_url(url: str) -> str:
    """
    Extract DOI from various publisher URL formats.

    Args:
        url: URL string that might contain a DOI

    Returns:
        Extracted DOI or None
    """
    if not url:
        return None

    url_lower = url.lower()

    # Direct DOI URLs
    if 'doi.org/' in url_lower:
        match = re.search(r'doi\.org/(10\.\d{4,}/[^\s?#]+)', url, re.IGNORECASE)
        if match:
            doi = match.group(1).rstrip('.,;)')
            return doi

    # Nature articles: nature.com/articles/s41560-019-0356-8 → 10.1038/s41560-019-0356-8
    if 'nature.com/articles/' in url_lower:
        match = re.search(r'nature\.com/articles/([^/?#]+)', url, re.IGNORECASE)
        if match:
            article_id = match.group(1).rstrip('.,;)')
            return f"10.1038/{article_id}"

    # MDPI: mdpi.com/2313-0105/8/10/151 → 10.3390/2313-0105/8/10/151
    if 'mdpi.com/' in url_lower:
        match = re.search(r'mdpi\.com/(\d{4}-\d{4}(?:/\d+)+)', url, re.IGNORECASE)
        if match:
            path = match.group(1).rstrip('.,;)')
            return f"10.3390/{path}"

    # IOP Science: iopscience.iop.org/article/10.1149/1945-7111/abae37 → 10.1149/1945-7111/abae37
    if 'iopscience.iop.org/article/' in url_lower:
        match = re.search(r'iopscience\.iop\.org/article/(10\.\d{4,}/[\w.-]+/[\w.-]+)', url, re.IGNORECASE)
        if match:
            doi = match.group(1).rstrip('.,;)')
            return doi

    # ScienceDirect PII: sciencedirect.com/science/article/pii/S2352152X24044748
    if 'sciencedirect.com/science/article/' in url_lower and '/pii/' in url_lower:
        match = re.search(r'/pii/([A-Z0-9]+)', url, re.IGNORECASE)
        if match:
            pii = match.group(1)
            # Try to look up DOI from PII via CrossRef
            doi = lookup_doi_from_pii(pii)
            if doi:
                return doi

    # Cell Press PII: cell.com/joule/fulltext/S2542-4351(24)00510-5
    if 'cell.com/' in url_lower and '/fulltext/' in url_lower:
        match = re.search(r'/fulltext/([A-Z0-9()-]+)', url, re.IGNORECASE)
        if match:
            pii = match.group(1).replace('(', '').replace(')', '')
            # Try to look up DOI from PII via CrossRef
            doi = lookup_doi_from_pii(pii)
            if doi:
                return doi

    # Wiley: onlinelibrary.wiley.com/doi/10.1002/adma.202402024 → 10.1002/adma.202402024
    if 'wiley.com/doi/' in url_lower:
        match = re.search(r'wiley\.com/doi/(?:full/|abs/)?(10\.\d{4,}/[^/?#]+)', url, re.IGNORECASE)
        if match:
            doi = match.group(1).rstrip('.,;)')
            return doi

    # Springer: link.springer.com/article/10.1007/s12274-024-6447-x → 10.1007/s12274-024-6447-x
    if 'springer.com/article/' in url_lower:
        match = re.search(r'springer\.com/article/(10\.\d{4,}/[^/?#]+)', url, re.IGNORECASE)
        if match:
            doi = match.group(1).rstrip('.,;)')
            return doi

    # Generic DOI pattern in URL (fallback)
    match = re.search(r'(10\.\d{4,}/[^\s?#]+)', url)
    if match:
        doi = match.group(1).rstrip('.,;)')
        return doi

    return None


def lookup_doi_from_pii(pii: str) -> str:
    """
    Look up DOI from PII (Publisher Item Identifier).

    For Elsevier papers (ScienceDirect, Cell), PII can often be converted to DOI:
    - Format: S + ISSN (8 chars) + year (2 digits) + serial (5-6 digits)
    - Example: S2352152X24044748 might map to specific DOI pattern

    However, there's no direct PII->DOI conversion API. Best approach:
    Return None and let Semantic Scholar handle it via title search.

    Args:
        pii: Publisher Item Identifier (e.g., S2352152X24044748)

    Returns:
        DOI if found, None otherwise
    """
    # PII to DOI lookup is unreliable without the publisher's internal database
    # Better to return None and fall back to Semantic Scholar title search
    # which has better coverage and accuracy
    return None


def find_doi_via_semantic_scholar(title: str, log_callback=None) -> str:
    """
    Find DOI by searching Semantic Scholar by paper title

    Args:
        title: Paper title to search for
        log_callback: Optional callback function for logging

    Returns:
        DOI string if found, empty string otherwise
    """
    import requests
    import time

    def log(msg):
        if log_callback:
            log_callback(msg)

    if not title:
        log("  [Semantic Scholar] Skipped: No title provided")
        return ''

    try:
        log(f"  [Semantic Scholar] Searching for: {title[:60]}...")

        # Semantic Scholar search API
        url = "https://api.semanticscholar.org/graph/v1/paper/search"
        params = {
            'query': title.strip(),
            'limit': 3,
            'fields': 'title,externalIds'
        }

        headers = {'User-Agent': 'AstrolabePaperDB/1.0'}

        response = requests.get(url, params=params, headers=headers, timeout=10)

        if response.status_code == 429:
            log("  [Semantic Scholar] Rate limited (429) - skipping")
            return ''

        response.raise_for_status()
        data = response.json()

        if 'data' not in data or not data['data']:
            log("  [Semantic Scholar] No results found")
            return ''

        log(f"  [Semantic Scholar] Found {len(data['data'])} result(s)")

        # Check top results for title match
        def normalize(s):
            return re.sub(r'[^\w\s]', '', s.lower()).strip()

        normalized_query = normalize(title)

        for i, paper in enumerate(data['data'], 1):
            result_title = paper.get('title', '')
            normalized_result = normalize(result_title)

            # Check for exact or very close match
            if normalized_query == normalized_result:
                log(f"  [Semantic Scholar] Match found: {result_title[:50]}...")

                external_ids = paper.get('externalIds', {})
                doi = external_ids.get('DOI', '')

                if doi:
                    log(f"  [Semantic Scholar] DOI: {doi}")
                    return doi
                else:
                    log(f"  [Semantic Scholar] No DOI in record")
                    return ''

        log("  [Semantic Scholar] No exact title match in results")
        return ''

    except requests.exceptions.Timeout:
        log("  [Semantic Scholar] Request timeout")
        return ''
    except requests.exceptions.RequestException as e:
        log(f"  [Semantic Scholar] Error: {type(e).__name__}")
        return ''
    except Exception as e:
        log(f"  [Semantic Scholar] Unexpected error: {type(e).__name__}")
        return ''


def normalize_title_for_matching(title: str) -> str:
    """Normalize title for duplicate detection."""
    if not title:
        return ""

    # Lowercase, remove punctuation, collapse whitespace
    normalized = title.lower()
    normalized = re.sub(r'[^\w\s]', '', normalized)
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    return normalized


def is_paper_in_library(title: str, doi: str, existing_papers: list) -> bool:
    """
    Check if paper is already in library.

    Args:
        title: Paper title
        doi: Paper DOI
        existing_papers: List of existing papers from library

    Returns:
        True if paper is duplicate, False otherwise
    """
    if not title:
        return False

    norm_title = normalize_title_for_matching(title)

    for paper in existing_papers:
        # Check by DOI if both have DOI
        if doi and paper.get('doi'):
            if doi.lower() == paper.get('doi', '').lower():
                return True

        # Check by title similarity
        paper_title = normalize_title_for_matching(paper.get('title', ''))
        if paper_title and norm_title:
            # Use simple similarity: check if 90% of words match
            title_words = set(norm_title.split())
            paper_words = set(paper_title.split())

            if title_words and paper_words:
                overlap = len(title_words & paper_words)
                similarity = overlap / max(len(title_words), len(paper_words))

                if similarity > 0.9:
                    return True

    return False


def get_column_value_case_insensitive(row_dict: dict, *possible_names: str) -> str:
    """
    Get value from dict with case-insensitive column name matching.

    Args:
        row_dict: Dictionary with column data
        possible_names: Possible column names to try

    Returns:
        Value from first matching column, or empty string
    """
    # Create lowercase mapping of all keys
    lowercase_map = {k.lower(): k for k in row_dict.keys()}

    # Try each possible name (case-insensitive)
    for name in possible_names:
        lowercase_name = name.lower()
        if lowercase_name in lowercase_map:
            actual_key = lowercase_map[lowercase_name]
            value = row_dict.get(actual_key, '')
            if value:
                return str(value).strip()

    return ''


# ===== CANONICAL IMPORT SCHEMA =====

# Define our canonical metadata fields
CANONICAL_SCHEMA = {
    'title': str,
    'authors': str,
    'year': str,
    'journal': str,
    'doi': str,
    'url': str,
    'abstract': str,
    'chemistry': str,
    'topics': str,
    'tags': str,
    'paper_type': str,
    'application': str,
    'pdf_status': str,
    'date_added': str,
    'notes': str
}

# Column mappings for different import sources
IMPORT_SOURCE_MAPPINGS = {
    'notion_csv': {
        'Title': 'title',
        'Authors / Orgs': 'authors',
        'Authors': 'authors',
        'Publication Year': 'year',
        'Year': 'year',
        'Journal': 'journal',
        'URL': 'url',
        'url': 'url',
        'Tags': 'tags',
        'tags': 'tags',
        'Abstract/Notes': 'abstract',
        'Abstract': 'abstract',
        'Notes': 'notes',
        'DOI': 'doi',
        'doi': 'doi'
    },
    'battery_datasets': {
        'title': 'title',
        'Title': 'title',
        'authors': 'authors',
        'Authors': 'authors',
        'year': 'year',
        'Year': 'year',
        'journal': 'journal',
        'Journal': 'journal',
        'paper_url': 'url',
        'chemistry': 'chemistry',
        'Chemistry': 'chemistry',
        'tags': 'tags',
        'Tags': 'tags',
        'doi': 'doi',
        'DOI': 'doi'
    },
    'generic': {
        # Fallback mappings - tries common variations
        'title': 'title',
        'Title': 'title',
        'TITLE': 'title',
        'authors': 'authors',
        'Authors': 'authors',
        'AUTHORS': 'authors',
        'Author': 'authors',
        'year': 'year',
        'Year': 'year',
        'YEAR': 'year',
        'journal': 'journal',
        'Journal': 'journal',
        'JOURNAL': 'journal',
        'url': 'url',
        'URL': 'url',
        'doi': 'doi',
        'DOI': 'doi',
        'abstract': 'abstract',
        'Abstract': 'abstract',
        'tags': 'tags',
        'Tags': 'tags',
        'chemistry': 'chemistry',
        'Chemistry': 'chemistry'
    }
}


def detect_import_source(columns: list) -> str:
    """
    Detect the import source type based on column names.

    Args:
        columns: List of column names from the file

    Returns:
        Source type: 'notion_csv', 'battery_datasets', or 'generic'
    """
    columns_lower = [c.lower() for c in columns]

    # Check for Battery Datasets signature
    if 'paper_url' in columns_lower or 'chemistry' in columns_lower:
        return 'battery_datasets'

    # Check for Notion signature
    if 'authors / orgs' in columns_lower or 'abstract/notes' in columns_lower:
        return 'notion_csv'

    # Default to generic
    return 'generic'


def normalize_to_canonical_schema(row_data: dict, source_type: str = None) -> dict:
    """
    Convert a row from any import source to our canonical schema.

    Args:
        row_data: Dictionary with source column names
        source_type: Type of source ('notion_csv', 'battery_datasets', 'generic')
                    If None, will auto-detect

    Returns:
        Dictionary with canonical field names
    """
    if source_type is None:
        source_type = detect_import_source(list(row_data.keys()))

    # Get the mapping for this source
    mapping = IMPORT_SOURCE_MAPPINGS.get(source_type, IMPORT_SOURCE_MAPPINGS['generic'])

    # Create canonical record
    canonical = {}

    # Map each source column to canonical field
    for source_col, canonical_field in mapping.items():
        if source_col in row_data and row_data[source_col]:
            value = str(row_data[source_col]).strip()
            if value:
                canonical[canonical_field] = value

    # Ensure all canonical fields exist (even if empty)
    for field in CANONICAL_SCHEMA.keys():
        if field not in canonical:
            canonical[field] = ''

    return canonical


def enrich_from_crossref(canonical_data: dict) -> dict:
    """
    Enrich missing fields using CrossRef API based on DOI or URL.

    Args:
        canonical_data: Dictionary with canonical fields

    Returns:
        Enriched dictionary with additional fields from CrossRef
    """
    # Extract DOI from URL if not present
    if not canonical_data['doi'] and canonical_data['url']:
        canonical_data['doi'] = extract_doi_from_url(canonical_data['url'])

    # If we have a DOI, try to fetch metadata
    if canonical_data['doi']:
        try:
            crossref_metadata = query_crossref_for_metadata(canonical_data['doi'])

            if crossref_metadata:
                # Fill in missing fields from CrossRef
                if not canonical_data['title'] and crossref_metadata.get('title'):
                    canonical_data['title'] = crossref_metadata['title']

                if not canonical_data['authors'] and crossref_metadata.get('authors'):
                    canonical_data['authors'] = ', '.join(crossref_metadata['authors'])

                if not canonical_data['year'] and crossref_metadata.get('year'):
                    canonical_data['year'] = str(crossref_metadata['year'])

                if not canonical_data['journal'] and crossref_metadata.get('journal'):
                    canonical_data['journal'] = crossref_metadata['journal']

                if not canonical_data['abstract'] and crossref_metadata.get('abstract'):
                    canonical_data['abstract'] = crossref_metadata['abstract']
        except Exception as e:
            # CrossRef enrichment failed, continue with what we have
            pass

    return canonical_data


def enrich_library_metadata(max_papers: int = None, progress_callback=None) -> dict:
    """
    Enrich metadata for papers in library that have URLs but missing key fields.

    Args:
        max_papers: Maximum number of papers to enrich (None = all)
        progress_callback: Optional callback function for progress updates

    Returns:
        Dict with enrichment statistics
    """
    import time

    # Load metadata
    metadata_file = Path("data/metadata.json")
    if not metadata_file.exists():
        return {'success': False, 'message': 'No metadata file found', 'enriched': 0}

    with open(metadata_file, 'r', encoding='utf-8') as f:
        all_metadata = json.load(f)

    # Find papers needing enrichment (have URL/DOI but missing key fields)
    papers_to_enrich = []
    for filename, paper in all_metadata.items():
        url = paper.get('url', '') or paper.get('source_url', '')
        doi = paper.get('doi', '')

        # Check if has URL or DOI
        if url or doi:
            # Check if missing key metadata
            missing_authors = not paper.get('authors') or paper.get('authors') == []
            missing_year = not paper.get('year')
            missing_journal = not paper.get('journal')

            if missing_authors or missing_year or missing_journal:
                papers_to_enrich.append((filename, paper, url, doi))

    if not papers_to_enrich:
        return {'success': True, 'message': 'No papers need enrichment', 'enriched': 0}

    # Limit if specified
    if max_papers:
        papers_to_enrich = papers_to_enrich[:max_papers]

    # Enrich papers
    enriched_count = 0
    failed_count = 0
    enrichment_logs = []  # Store logs for debugging

    def log(msg):
        """Helper to log messages"""
        enrichment_logs.append(msg)
        if progress_callback:
            progress_callback(-1, -1, msg)  # Special callback for logs

    for idx, (filename, paper, url, doi) in enumerate(papers_to_enrich):
        try:
            title = paper.get('title', filename)
            if progress_callback:
                progress_callback(idx + 1, len(papers_to_enrich), title)

            log(f"\n[{idx + 1}/{len(papers_to_enrich)}] {title[:60]}...")

            # Step 1: Extract DOI from URL if not present
            if not doi and url:
                log(f"  [DOI Extraction] URL: {url[:70]}...")
                doi = extract_doi_from_url(url)
                if doi:
                    log(f"  [DOI Extraction] Found: {doi}")
                    all_metadata[filename]['doi'] = doi
                else:
                    log(f"  [DOI Extraction] Failed: No DOI pattern in URL")

            # Step 2: If no DOI, try Semantic Scholar as fallback
            if not doi:
                log(f"  [Fallback] Trying Semantic Scholar by title...")
                doi = find_doi_via_semantic_scholar(title, log_callback=log)
                if doi:
                    all_metadata[filename]['doi'] = doi
                    # Rate limit for Semantic Scholar API (100 requests per 5 min)
                    time.sleep(1.5)

            # Step 3: Query CrossRef if we have DOI
            if doi:
                log(f"  [CrossRef] Querying for DOI: {doi}")
                crossref_metadata = query_crossref_for_metadata(doi)

                if crossref_metadata:
                    log(f"  [CrossRef] Received metadata with {len(crossref_metadata)} fields")
                    updated_fields = []

                    # Update missing fields
                    if not all_metadata[filename].get('authors') or all_metadata[filename].get('authors') == []:
                        if crossref_metadata.get('authors'):
                            all_metadata[filename]['authors'] = crossref_metadata['authors']
                            updated_fields.append('authors')

                    if not all_metadata[filename].get('year'):
                        if crossref_metadata.get('year'):
                            all_metadata[filename]['year'] = str(crossref_metadata['year'])
                            updated_fields.append('year')

                    if not all_metadata[filename].get('journal'):
                        if crossref_metadata.get('journal'):
                            all_metadata[filename]['journal'] = crossref_metadata['journal']
                            updated_fields.append('journal')

                    if not all_metadata[filename].get('abstract'):
                        if crossref_metadata.get('abstract'):
                            all_metadata[filename]['abstract'] = crossref_metadata['abstract']
                            updated_fields.append('abstract')

                    if not all_metadata[filename].get('volume'):
                        if crossref_metadata.get('volume'):
                            all_metadata[filename]['volume'] = crossref_metadata['volume']
                            updated_fields.append('volume')

                    if not all_metadata[filename].get('issue'):
                        if crossref_metadata.get('issue'):
                            all_metadata[filename]['issue'] = crossref_metadata['issue']
                            updated_fields.append('issue')

                    if not all_metadata[filename].get('pages'):
                        if crossref_metadata.get('pages'):
                            all_metadata[filename]['pages'] = crossref_metadata['pages']
                            updated_fields.append('pages')

                    if updated_fields:
                        log(f"  [Success] Updated: {', '.join(updated_fields)}")
                        enriched_count += 1
                    else:
                        log(f"  [Skipped] All fields already present")
                        failed_count += 1
                else:
                    log(f"  [CrossRef] Failed: No metadata returned")
                    failed_count += 1
            else:
                log(f"  [Failed] No DOI available - cannot enrich")
                failed_count += 1

            # Rate limiting - 1 second delay between papers
            time.sleep(1.0)

        except Exception as e:
            log(f"  [Error] {type(e).__name__}: {str(e)}")
            failed_count += 1

    # Save updated metadata to metadata.json
    with open(metadata_file, 'w', encoding='utf-8') as f:
        json.dump(all_metadata, f, indent=2, ensure_ascii=False)

    # Update ChromaDB with enriched metadata
    from lib.rag import DatabaseClient
    DatabaseClient.clear_cache()
    collection = DatabaseClient.get_collection()

    # Update each enriched paper in ChromaDB
    for filename, paper, url, doi in papers_to_enrich:
        enriched_paper = all_metadata[filename]
        doc_id = f"{filename}_metadata"

        # Get existing metadata document
        try:
            existing_doc = collection.get(ids=[doc_id], include=['metadatas'])
            if existing_doc and existing_doc['metadatas']:
                # Update the metadata document with enriched fields
                authors = enriched_paper.get('authors', [])
                if isinstance(authors, list):
                    authors_str = '; '.join(authors)
                else:
                    authors_str = str(authors)

                # Delete old and add updated
                collection.delete(ids=[doc_id])
                collection.add(
                    documents=[f"Metadata: {enriched_paper.get('title', '')}. DOI: {enriched_paper.get('doi', '')}"],
                    metadatas=[{
                        'filename': filename,
                        'page_num': 0,
                        'paper_type': enriched_paper.get('paper_type', 'experimental'),
                        'application': enriched_paper.get('application', 'general'),
                        'chemistries': ','.join(enriched_paper.get('chemistries', [])) if isinstance(enriched_paper.get('chemistries'), list) else '',
                        'topics': ','.join(enriched_paper.get('topics', [])) if isinstance(enriched_paper.get('topics'), list) else '',
                        'section_name': 'metadata',
                        'abstract': enriched_paper.get('abstract', ''),
                        'author_keywords': ';'.join(enriched_paper.get('author_keywords', [])) if isinstance(enriched_paper.get('author_keywords'), list) else '',
                        'title': enriched_paper.get('title', ''),
                        'authors': authors_str,
                        'year': str(enriched_paper.get('year', '')),
                        'journal': enriched_paper.get('journal', ''),
                        'doi': enriched_paper.get('doi', '')
                    }],
                    ids=[doc_id]
                )
        except:
            pass  # If document doesn't exist in ChromaDB, skip

    # Clear cache again after updates
    DatabaseClient.clear_cache()

    # Create appropriate message
    if enriched_count == 0 and failed_count == 0:
        message = "No papers were enriched"
    elif enriched_count == 0:
        message = f"Failed to enrich {failed_count} paper(s)"
    elif failed_count == 0:
        message = f"Successfully enriched {enriched_count} paper(s)"
    else:
        message = f"Enriched {enriched_count} paper(s), failed {failed_count}"

    return {
        'success': True,
        'enriched': enriched_count,
        'failed': failed_count,
        'total': len(papers_to_enrich),
        'message': message,
        'logs': enrichment_logs  # Include detailed logs for debugging
    }


def import_csv_papers(csv_papers: list, batch_size: int, skip_existing: bool, existing_papers: list):
    """
    Import papers from CSV with progress tracking and rate limiting.

    Args:
        csv_papers: List of paper dicts from CSV
        batch_size: Number of papers to import
        skip_existing: Whether to skip duplicates
        existing_papers: List of existing papers in library
    """
    import time

    # Debug: Show column names and detected source type
    if csv_papers:
        columns = list(csv_papers[0].keys())
        source_type = detect_import_source(columns)

        st.info(f"📋 **Import Source Detected:** {source_type.replace('_', ' ').title()}")
        st.caption(f"**Columns:** {', '.join(columns)}")

        # Show field mapping
        with st.expander("🔍 Column Mapping Preview"):
            sample = normalize_to_canonical_schema(csv_papers[0], source_type)
            mapped_fields = {k: v for k, v in sample.items() if v}
            if mapped_fields:
                st.json(mapped_fields)

    # Progress tracking
    progress_bar = st.progress(0)
    status_text = st.empty()

    imported = 0
    skipped = 0
    failed = 0
    import_logs = []  # Collect detailed logs

    # Process papers in batch
    papers_to_process = csv_papers[:batch_size]

    for idx, csv_paper in enumerate(papers_to_process):
        try:
            # Update progress
            progress = (idx + 1) / len(papers_to_process)
            progress_bar.progress(progress)

            # Normalize to canonical schema
            canonical = normalize_to_canonical_schema(csv_paper)

            # Enrich with CrossRef if we have URL/DOI
            canonical = enrich_from_crossref(canonical)

            # Extract fields
            title = canonical['title']
            url = canonical['url']
            authors = canonical['authors']
            year = canonical['year']
            journal = canonical['journal']
            abstract = canonical['abstract']
            tags = canonical['tags']
            doi = canonical['doi']
            chemistry = canonical['chemistry']

            # Skip if no title
            if not title:
                import_logs.append(f"⚠️ Row {idx + 1}: No title, skipping")
                skipped += 1
                time.sleep(0.5)
                continue

            # Check for duplicates
            if skip_existing and is_paper_in_library(title, doi, existing_papers):
                status_text.text(f"Importing paper {idx + 1} of {len(papers_to_process)}: {title[:50]}...")
                import_logs.append(f"⏭️ Skipped: {title[:60]}... (already in library)")
                skipped += 1
                time.sleep(0.5)
                continue

            # Show current paper
            status_text.text(f"Importing paper {idx + 1} of {len(papers_to_process)}: {title[:50]}...")

            result = add_paper_with_pdf_search(
                doi=doi or '',
                title=title,
                authors=authors,
                year=year,
                url=url
            )

            if result['success']:
                pdf_icon = "📄" if result['pdf_found'] else "📝"
                import_logs.append(f"✓ {pdf_icon} Added: {title[:60]}...")
                imported += 1
            else:
                import_logs.append(f"❌ Failed: {title[:60]}... - {result['message']}")
                failed += 1

            # Rate limiting: 2-3 seconds between papers (for CrossRef + Unpaywall)
            time.sleep(2.5)

        except Exception as e:
            import traceback
            error_msg = f"❌ Error processing row {idx + 1}: {type(e).__name__}: {str(e)}"
            import_logs.append(error_msg)
            import_logs.append(f"   Traceback: {traceback.format_exc()}")
            failed += 1
            time.sleep(1)

    # Clear progress indicators
    progress_bar.empty()
    status_text.empty()

    # Summary message
    summary_parts = []
    if imported > 0:
        summary_parts.append(f"✅ Imported {imported} paper{'s' if imported != 1 else ''}")
    if skipped > 0:
        summary_parts.append(f"⏭️ {skipped} skipped")
    if failed > 0:
        summary_parts.append(f"❌ {failed} failed")

    if imported > 0:
        st.success(". ".join(summary_parts) + ".")
    elif len(papers_to_process) > 0:
        st.info(". ".join(summary_parts) + ".")
    else:
        st.warning("No papers to import")

    # Details in expander
    if import_logs:
        with st.expander("Show details"):
            for log in import_logs:
                st.text(log)

    if imported > 0:
        st.info("Refreshing app to show new papers...")
        time.sleep(2)
        st.rerun()


def update_paper_metadata(filename: str, doi: str, crossref_metadata: dict):
    """Update paper metadata in metadata.json and ChromaDB."""
    metadata_file = Path("data/metadata.json")

    # Load existing metadata
    if metadata_file.exists():
        with open(metadata_file, 'r', encoding='utf-8') as f:
            all_metadata = json.load(f)
    else:
        all_metadata = {}

    # Update metadata for this paper
    if filename in all_metadata:
        # Update with CrossRef data
        all_metadata[filename]['doi'] = doi
        if crossref_metadata:
            all_metadata[filename]['title'] = crossref_metadata.get('title', all_metadata[filename].get('title', ''))
            all_metadata[filename]['authors'] = crossref_metadata.get('authors', all_metadata[filename].get('authors', []))
            all_metadata[filename]['year'] = crossref_metadata.get('year', all_metadata[filename].get('year', ''))
            all_metadata[filename]['journal'] = crossref_metadata.get('journal', all_metadata[filename].get('journal', ''))

        # Save updated metadata
        with open(metadata_file, 'w', encoding='utf-8') as f:
            json.dump(all_metadata, f, indent=2, ensure_ascii=False)

        # Update ChromaDB
        update_chromadb_metadata(filename, all_metadata[filename])

        return True
    return False


def update_chromadb_metadata(filename: str, paper_metadata: dict):
    """Update metadata for all chunks of a paper in ChromaDB."""
    try:
        collection = rag.DatabaseClient.get_collection()

        # Get all chunks for this paper
        results = collection.get(
            where={"filename": filename},
            include=["metadatas"]
        )

        if not results['ids']:
            return

        # Update metadata for each chunk
        updated_metadatas = []
        for metadata in results['metadatas']:
            metadata['doi'] = paper_metadata.get('doi', '')
            metadata['title'] = paper_metadata.get('title', '')
            metadata['authors'] = ';'.join(paper_metadata.get('authors', []))
            metadata['year'] = paper_metadata.get('year', '')
            metadata['journal'] = paper_metadata.get('journal', '')
            updated_metadatas.append(metadata)

        # Update in ChromaDB
        collection.update(
            ids=results['ids'],
            metadatas=updated_metadatas
        )
    except Exception as e:
        st.error(f"Error updating ChromaDB: {e}")


def main():
    # Import modules used throughout main() to avoid UnboundLocalError
    import re
    from pathlib import Path
    import time as timing_module
    import sys

    # TIMING: Start
    _start_time = timing_module.time()
    print(f"\n[TIMING] main() started", file=sys.stderr, flush=True)

    # Initialize session state
    if "selected_paper" not in st.session_state:
        st.session_state.selected_paper = None
    if "query_result" not in st.session_state:
        st.session_state.query_result = None
    if "active_tab" not in st.session_state:
        st.session_state.active_tab = "Library"

    # Initialize theme
    if "theme" not in st.session_state:
        st.session_state.theme = load_theme_preference()

    # Initialize databases
    read_status.init_db()
    query_history.init_db()
    collections._get_connection().close()  # Initialize collections DB

    # Apply professional CSS styling
    current_theme = st.session_state.theme
    st.markdown(styles.get_professional_css(current_theme), unsafe_allow_html=True)

    # Header - Clean and professional
    st.markdown("""
        <div class="app-header">
            <h1 class="app-title">Astrolabe Research Library</h1>
            <p class="app-subtitle">Battery research papers with AI-powered search</p>
        </div>
    """, unsafe_allow_html=True)

    # Load resources using backend with session state caching
    # Cache papers to avoid reloading ChromaDB on every rerun
    if 'cached_papers' not in st.session_state or st.session_state.get('reload_papers', False):
        print(f"[TIMING] RELOADING papers from ChromaDB (cache miss or reload_papers={st.session_state.get('reload_papers', 'not set')})", file=sys.stderr, flush=True)
        try:
            _reload_start = timing_module.time()
            st.session_state.cached_papers = rag.get_paper_library()
            st.session_state.cached_filter_options = rag.get_filter_options()
            st.session_state.cached_total_chunks = rag.get_collection_count()
            st.session_state.reload_papers = False
            print(f"[TIMING] Papers reloaded in {timing_module.time() - _reload_start:.3f}s", file=sys.stderr, flush=True)
        except (FileNotFoundError, RuntimeError) as e:
            st.error(str(e))
            st.info("Please run `python scripts/ingest.py` first to create the database")
            st.stop()
    else:
        print(f"[TIMING] Using cached papers (no reload)", file=sys.stderr, flush=True)

    papers = st.session_state.cached_papers
    filter_options = st.session_state.cached_filter_options
    total_chunks = st.session_state.cached_total_chunks

    # TIMING: After loading papers
    print(f"[TIMING] Papers loaded: {timing_module.time() - _start_time:.3f}s", file=sys.stderr, flush=True)

    # Sidebar - Simplified and professional
    with st.sidebar:
        # Quick stats with breakdown
        st.subheader("Library Stats")

        # Calculate stats (cached in session state to avoid slow disk I/O on every rerun)
        if 'cached_stats' not in st.session_state or st.session_state.get('reload_papers', False):
            print(f"[TIMING] RECALCULATING sidebar stats (checking {len(papers)} PDFs)", file=sys.stderr, flush=True)
            _stats_start = timing_module.time()
            total_papers = len(papers)
            complete_papers = 0
            metadata_only_papers = 0
            incomplete_papers = 0

            for paper in papers:
                # Check if metadata is complete
                has_title = bool(paper.get('title', '').strip())
                has_authors = bool(paper.get('authors') and paper.get('authors') != [])
                has_year = bool(paper.get('year', '').strip())
                has_journal = bool(paper.get('journal', '').strip())

                metadata_complete = has_title and has_authors and has_year and has_journal

                # Check if PDF exists
                filename = paper.get('filename', '')
                has_pdf = False
                if filename:
                    pdf_path = Path("papers") / filename
                    has_pdf = pdf_path.exists()

                # Categorize
                if metadata_complete and has_pdf:
                    complete_papers += 1
                elif metadata_complete and not has_pdf:
                    metadata_only_papers += 1
                else:
                    incomplete_papers += 1

            st.session_state.cached_stats = {
                'total': total_papers,
                'complete': complete_papers,
                'metadata_only': metadata_only_papers,
                'incomplete': incomplete_papers
            }
            print(f"[TIMING] Stats calculated in {timing_module.time() - _stats_start:.3f}s", file=sys.stderr, flush=True)
        else:
            # Use cached stats
            print(f"[TIMING] Using cached stats (no recalc)", file=sys.stderr, flush=True)
            total_papers = st.session_state.cached_stats['total']
            complete_papers = st.session_state.cached_stats['complete']
            metadata_only_papers = st.session_state.cached_stats['metadata_only']
            incomplete_papers = st.session_state.cached_stats['incomplete']

        # Display total with breakdown
        st.metric("Total Papers", total_papers)

        # Calculate percentages
        complete_pct = (complete_papers / total_papers * 100) if total_papers > 0 else 0
        metadata_pct = (metadata_only_papers / total_papers * 100) if total_papers > 0 else 0
        incomplete_pct = (incomplete_papers / total_papers * 100) if total_papers > 0 else 0

        # Compact summary text
        st.caption(
            f"{complete_papers} complete ({complete_pct:.0f}%) | "
            f"{metadata_only_papers} metadata only ({metadata_pct:.0f}%) | "
            f"{incomplete_papers} incomplete ({incomplete_pct:.0f}%)"
        )

        # Visual progress bars
        st.caption("**Data Coverage**")
        st.progress(complete_pct / 100, text=f"✅ Complete: {complete_papers}")
        st.progress(metadata_pct / 100, text=f"📋 Metadata Only: {metadata_only_papers}")
        st.progress(incomplete_pct / 100, text=f"⚠️ Incomplete: {incomplete_papers}")

        st.divider()
        st.metric("Chunks", total_chunks)

    # TIMING: After sidebar
    print(f"[TIMING] Sidebar rendered: {timing_module.time() - _start_time:.3f}s", file=sys.stderr, flush=True)

    # Main content - Tabs
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(["📥 Import", "Library", "🔍 Discover", "Research", "History", "Settings", "🗑️ Trash"])

    # ===== TAB 1: IMPORT =====
    with tab1:
        st.session_state.active_tab = "Import"

        st.markdown("### Import Papers")
        st.caption("Add papers to your library using one of the methods below")

        st.markdown("---")

        # Section 1: Add by URL or DOI
        st.subheader("📝 Add by URL or DOI")
        st.caption("Paste a DOI or URL to fetch metadata and search for open access PDF")

        with st.form("import_doi_form", clear_on_submit=True):
            doi_url_input = st.text_input(
                "DOI or URL:",
                placeholder="10.1016/j.jpowsour.2024.234567 or https://doi.org/...",
                help="Enter a DOI (e.g., 10.1016/...) or URL to a paper"
            )

            submit_doi = st.form_submit_button("📥 Add to Library", type="primary", use_container_width=True)

            if submit_doi and doi_url_input.strip():
                with st.spinner("Adding paper..."):
                    input_text = doi_url_input.strip()

                    # Extract DOI from input (handle URLs)
                    doi = None
                    if input_text.startswith('http'):
                        # Try to extract DOI from URL
                        doi_match = re.search(r'10\.\d{4,}/[^\s]+', input_text)
                        if doi_match:
                            doi = doi_match.group(0)
                    else:
                        # Assume it's a DOI
                        doi = input_text

                    if doi:
                        # Get metadata from CrossRef
                        crossref_metadata = query_crossref_for_metadata(doi)

                        if crossref_metadata and crossref_metadata.get('title'):
                            # Use add_paper_with_pdf_search to add with PDF search
                            result = add_paper_with_pdf_search(
                                doi=doi,
                                title=crossref_metadata.get('title', ''),
                                authors=', '.join(crossref_metadata.get('authors', [])),
                                year=crossref_metadata.get('year', ''),
                                url=input_text
                            )

                            if result['success']:
                                if result['pdf_found']:
                                    st.success("✓ Paper added with PDF!")
                                else:
                                    st.success("✓ Paper added (metadata-only)")
                                    st.caption("No open access PDF found")

                                time.sleep(1)
                                st.rerun()
                            else:
                                st.error(f"Failed: {result['message']}")
                        else:
                            st.error("Could not fetch metadata from CrossRef")
                    else:
                        st.error("Invalid DOI or URL format")

        st.markdown("---")

        # Section 2: Upload PDFs
        st.subheader("📂 Upload PDFs")
        st.caption("Drag and drop PDF files or click to browse")

        uploaded_files = st.file_uploader(
            "Upload PDFs",
            type=['pdf'],
            accept_multiple_files=True,
            label_visibility="collapsed",
            help="Upload one or more PDF files to add to your library"
        )

        if uploaded_files:
            st.info(f"📄 {len(uploaded_files)} file(s) selected")

            if st.button("Upload and Process", type="primary", use_container_width=True):
                with st.spinner(f"Uploading {len(uploaded_files)} file(s)..."):
                    papers_dir = Path("papers")
                    papers_dir.mkdir(exist_ok=True)

                    success_count = 0
                    for uploaded_file in uploaded_files:
                        try:
                            # Save to papers/ directory
                            file_path = papers_dir / uploaded_file.name
                            with open(file_path, 'wb') as f:
                                f.write(uploaded_file.getbuffer())
                            success_count += 1
                        except Exception as e:
                            st.error(f"Failed to upload {uploaded_file.name}: {str(e)}")

                    if success_count > 0:
                        st.success(f"✓ Uploaded {success_count} file(s) to papers/ folder")
                        st.info("Run the ingestion pipeline to process them")
                        time.sleep(2)
                        st.rerun()

        st.markdown("---")

        # Section 3: Scan Papers Folder
        st.subheader("🔍 Scan Papers Folder")
        st.caption("Check the papers/ folder for new PDF files not yet in the library")

        if st.button("📂 Scan for New PDFs", type="primary", use_container_width=True):
            with st.spinner("Scanning papers/ folder..."):
                # Get list of PDF files in papers/ folder
                papers_dir = Path("papers")
                if papers_dir.exists():
                    pdf_files = list(papers_dir.glob("*.pdf"))

                    # Load metadata to check which are new
                    metadata_file = Path("data/metadata.json")
                    existing_filenames = set()
                    if metadata_file.exists():
                        with open(metadata_file, 'r', encoding='utf-8') as f:
                            metadata = json.load(f)
                            existing_filenames = set(metadata.keys())

                    # Find new PDFs
                    new_pdfs = [pdf for pdf in pdf_files if pdf.name not in existing_filenames]

                    if new_pdfs:
                        st.success(f"✓ Found {len(new_pdfs)} new PDF(s)")
                        st.caption("Run ingestion pipeline to process them:")
                        st.code("python scripts/ingest_pipeline.py", language="bash")
                        st.markdown("**New files:**")
                        for pdf in new_pdfs[:10]:  # Show first 10
                            st.caption(f"• {pdf.name}")
                        if len(new_pdfs) > 10:
                            st.caption(f"... and {len(new_pdfs) - 10} more")
                    else:
                        st.info("No new PDFs found - all files in papers/ folder are already in the library")
                else:
                    st.warning("papers/ folder not found")

        st.markdown("---")

        # Section 4: Import from CSV/Excel
        st.subheader("📊 Import from CSV/Excel")
        st.caption("Bulk import papers from CSV or Excel files (Notion exports, Battery Datasets catalog, etc.)")

        with st.expander("ℹ️ Supported Formats & Instructions", expanded=False):
            st.markdown("""
            **Supported File Types:**
            - CSV files (.csv) - Notion exports, custom lists
            - Excel files (.xlsx, .xls) - Battery Datasets catalog, spreadsheets

            **CSV Format (Notion, etc.):**
            - `Title` - Paper title (required)
            - `URL` - DOI or paper URL (for metadata fetching)
            - `Authors / Orgs` or `Authors` - Author names
            - `Publication Year` or `Year` - Publication year
            - `Journal` - Journal name
            - `Tags` - Keywords/tags (comma-separated)
            - `Abstract/Notes` or `Abstract` - Paper abstract

            **Excel Format (Battery Datasets):**
            - Sheet: "Battery_Datasets" (auto-detected)
            - `paper_url` - URL with DOI (for metadata fetching)
            - `Title` or `title` - Paper title (required)
            - `chemistry` - Battery chemistry tags
            - `authors` - Author names
            - `journal` - Journal name
            - `tags` - Additional tags

            **Import Process:**
            1. Upload your CSV or Excel file
            2. Preview papers to be imported
            3. Set batch size (start small to test)
            4. Click Import to process papers
            5. For each paper with a URL/DOI:
               - Extracts DOI from URL
               - Fetches full metadata from CrossRef
               - Searches for open access PDF (Unpaywall)
               - Downloads PDF if available
            6. Skips papers already in library
            7. Respects API rate limits (2-3 sec between requests)

            **Tips:**
            - Start with batch size of 5-10 to test
            - Papers without DOI will use file metadata only
            - Duplicate detection based on DOI + title matching
            """)

        upload_file = st.file_uploader(
            "Upload CSV or Excel file",
            type=['csv', 'xlsx', 'xls'],
            help="Upload a CSV or Excel file with paper metadata"
        )

        if upload_file is not None:
            try:
                import csv
                import io

                file_papers = []
                file_type = upload_file.name.split('.')[-1].lower()

                # Read CSV
                if file_type == 'csv':
                    csv_content = upload_file.read().decode('utf-8')
                    csv_reader = csv.DictReader(io.StringIO(csv_content))
                    file_papers = list(csv_reader)
                    st.success(f"✓ Loaded {len(file_papers)} papers from CSV")

                # Read Excel
                elif file_type in ['xlsx', 'xls']:
                    import openpyxl

                    # Read Excel file
                    workbook = openpyxl.load_workbook(upload_file)

                    # Try to find Battery_Datasets sheet, otherwise use first sheet
                    if 'Battery_Datasets' in workbook.sheetnames:
                        sheet = workbook['Battery_Datasets']
                        st.info("📊 Detected Battery Datasets catalog format")
                    else:
                        sheet = workbook.active
                        st.info(f"📊 Using sheet: {sheet.title}")

                    # Read headers
                    headers = []
                    for cell in sheet[1]:
                        headers.append(cell.value)

                    # Read data rows
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if any(row):  # Skip empty rows
                            paper_dict = {}
                            for i, value in enumerate(row):
                                if i < len(headers) and headers[i]:
                                    paper_dict[headers[i]] = str(value) if value is not None else ''

                            # Map Battery Datasets columns to standard format
                            if 'paper_url' in paper_dict:
                                paper_dict['URL'] = paper_dict.get('paper_url', '')
                            if 'chemistry' in paper_dict:
                                # Merge chemistry with tags
                                existing_tags = paper_dict.get('tags', paper_dict.get('Tags', ''))
                                chemistry_tag = paper_dict.get('chemistry', '')
                                if chemistry_tag and existing_tags:
                                    paper_dict['Tags'] = f"{existing_tags}, {chemistry_tag}"
                                elif chemistry_tag:
                                    paper_dict['Tags'] = chemistry_tag

                            # Ensure Title field exists
                            if 'title' in paper_dict and 'Title' not in paper_dict:
                                paper_dict['Title'] = paper_dict['title']

                            file_papers.append(paper_dict)

                    st.success(f"✓ Loaded {len(file_papers)} papers from Excel")

                csv_papers = file_papers  # Use common variable name for rest of code

                st.success(f"✓ Ready to import {len(csv_papers)} papers")

                # Show preview
                with st.expander("📋 Preview CSV Data (first 5 rows)", expanded=True):
                    preview_df = pd.DataFrame(csv_papers[:5])
                    st.dataframe(preview_df, use_container_width=True)

                # Batch size selector
                st.markdown("**Import Settings**")
                col_batch, col_skip = st.columns(2)

                with col_batch:
                    batch_size = st.number_input(
                        "Batch size",
                        min_value=1,
                        max_value=100,
                        value=10,
                        help="Number of papers to import (start small to test)"
                    )

                with col_skip:
                    skip_existing = st.checkbox(
                        "Skip duplicates",
                        value=True,
                        help="Skip papers already in library"
                    )

                st.caption(f"Will import up to {min(batch_size, len(csv_papers))} papers")

                # Import button
                if st.button("📥 Import Papers", type="primary", use_container_width=True):
                    try:
                        import traceback
                        import_csv_papers(csv_papers, batch_size, skip_existing, papers)
                    except Exception as e:
                        st.error("❌ **Import Failed - Full Error Details:**")
                        st.error(traceback.format_exc())
                        st.markdown("---")
                        st.error(f"**Error Type:** {type(e).__name__}")
                        st.error(f"**Error Message:** {str(e)}")

            except Exception as e:
                st.error(f"Error reading file: {str(e)}")
                st.caption("Make sure your file is properly formatted")
                import traceback
                with st.expander("📋 Error Details"):
                    st.code(traceback.format_exc())

    # ===== TAB 2: LIBRARY =====
    with tab2:
        st.session_state.active_tab = "Library"

        # Upload section (only show when not viewing a paper detail)
        if not st.session_state.selected_paper:
            st.markdown("### Add Papers to Library")

            # Side-by-side layout for import options
            col_left, col_right = st.columns(2)

            with col_left:
                # URL import section
                with st.expander("Import from URL or DOI", expanded=False):
                    # Hide the "Press Enter to submit" message
                    st.markdown("""
                        <style>
                        /* Hide form submission instructions */
                        .stForm [data-testid="stMarkdownContainer"] p:contains("Press Enter") {
                            display: none !important;
                        }
                        .stForm small {
                            display: none !important;
                        }
                        [data-testid="InputInstructions"] {
                            display: none !important;
                        }
                        </style>
                    """, unsafe_allow_html=True)

                    with st.form("url_import_form", clear_on_submit=False):
                        st.caption("**Import from URL**")
                        url_input = st.text_input(
                            "Paste URL here",
                            placeholder="https://arxiv.org/abs/... or https://ieeexplore.ieee.org/...",
                            label_visibility="collapsed",
                            key="url_import_input"
                        )

                        st.caption("**Or enter DOI directly**")
                        doi_input = st.text_input(
                            "Enter DOI",
                            placeholder="10.1016/j.jpowsour.2024.235555",
                            label_visibility="collapsed",
                            key="doi_import_input"
                        )

                        st.caption("**Supported formats:**")
                        st.caption("• arXiv: `arxiv.org/abs/...` or `arxiv.org/pdf/...`")
                        st.caption("• DOI: `doi.org/10.xxxx/...` or `10.xxxx/...`")
                        st.caption("• Publisher pages: IEEE, ScienceDirect, Wiley, Springer, Nature, MDPI, ACS, RSC, IOP, etc.")
                        st.caption("• Direct PDF: Any URL ending in `.pdf`")

                        st.warning("⚠️ **Publisher Blocking:** Many publishers (especially ScienceDirect, Wiley, Springer) block automated access to their article pages. If you get a \"403 Forbidden\" error, use the **DOI field** instead, which bypasses the publisher page entirely.")

                        # Determine what to import
                        import_input = None
                        if url_input and doi_input:
                            st.warning("⚠️ Please use either URL or DOI, not both")
                        elif doi_input:
                            # Treat DOI as a doi.org URL
                            doi_clean = doi_input.strip()
                            if not doi_clean.startswith('http'):
                                import_input = f"https://doi.org/{doi_clean}"
                            else:
                                import_input = doi_clean
                        elif url_input:
                            import_input = url_input

                        submit_button = st.form_submit_button("Import", type="primary", use_container_width=True)

                    # Process result outside the form
                    if submit_button:
                        if not import_input:
                            st.warning("⚠️ Please enter a URL or DOI")

                    if submit_button and import_input:
                        # Create a container for progress updates
                        progress_container = st.container()

                        # Process the URL or DOI
                        result = process_url_import(import_input, progress_container)

                        # Show results
                        if result['success']:
                            if result['metadata_only']:
                                st.success(f"✅ Metadata saved for: **{result['title']}**")
                                st.info(f"📄 Filename: {result['filename']}")
                                st.info("📌 No open access PDF available. You can manually upload the PDF later.")
                                time.sleep(3)
                                st.rerun()
                            else:
                                st.success(f"✅ Successfully imported: **{result['title'] or result['filename']}**")
                                st.balloons()
                                time.sleep(2)
                                st.rerun()
                        else:
                            st.error(f"❌ Import failed: {result['error']}")

            with col_right:
                # Drag-and-drop PDF upload
                uploaded_files = st.file_uploader(
                    "Upload PDFs",
                    type=['pdf'],
                    accept_multiple_files=True,
                    label_visibility="collapsed",
                    key="pdf_uploader"
                )

                if uploaded_files:

                    # Check for duplicates
                    papers_dir = Path("papers")
                    papers_dir.mkdir(parents=True, exist_ok=True)

                    duplicates = []
                    new_files = []
                    for file in uploaded_files:
                        if (papers_dir / file.name).exists():
                            duplicates.append(file)
                        else:
                            new_files.append(file)

                    # Show what was uploaded
                    st.write(f"**{len(uploaded_files)} file(s) selected:**")

                    if new_files:
                        st.success(f"✅ {len(new_files)} new file(s):")
                        for file in new_files:
                            st.write(f"- {file.name} ({file.size / 1024:.1f} KB)")

                    if duplicates:
                        st.warning(f"⚠️ {len(duplicates)} duplicate(s) found:")
                        for file in duplicates:
                            st.write(f"- {file.name} (already exists in library)")

                        duplicate_action = st.radio(
                            "How to handle duplicates?",
                            ["Skip duplicates", "Replace existing papers"],
                            key="duplicate_action"
                        )
                    else:
                        duplicate_action = "Skip duplicates"

                    col1, col2 = st.columns([1, 4])
                    with col1:
                        process_disabled = len(new_files) == 0 and duplicate_action == "Skip duplicates"
                        if st.button("🚀 Process Papers", type="primary", disabled=process_disabled):
                            # Create a container for progress updates
                            progress_container = st.container()

                            # Process the files
                            replace_mode = duplicate_action == "Replace existing papers"
                            results = process_uploaded_pdfs(uploaded_files, progress_container, replace_duplicates=replace_mode)

                            # Show results
                            total_processed = len(results['saved']) + len(results['replaced'])
                            if total_processed > 0:
                                if results['replaced']:
                                    st.success(f"✅ Successfully processed {total_processed} paper(s)!")
                                    st.info(f"📝 Replaced {len(results['replaced'])} existing paper(s)")
                                else:
                                    st.success(f"✅ Successfully added {len(results['saved'])} paper(s) to library!")

                                # Invalidate cache after adding papers
                                st.session_state.reload_papers = True
                                st.toast(f"Processed {total_processed} paper(s)", icon="✅")

                            if results['skipped']:
                                st.warning(f"⚠️ Skipped {len(results['skipped'])} duplicate(s):")
                                for filename in results['skipped']:
                                    st.caption(f"- {filename} (already exists)")

                            if results['failed']:
                                st.error(f"❌ Failed to process {len(results['failed'])} file(s):")
                                for filename, error in results['failed']:
                                    st.caption(f"- {filename}: {error}")

                            # Clear the uploader and reload
                            if total_processed > 0:
                                st.balloons()
                                time.sleep(2)
                                st.rerun()

                    with col2:
                        if st.button("Cancel"):
                            st.rerun()

            st.divider()

        if st.session_state.selected_paper:
            # Detail view - Clean layout
            paper_filename = st.session_state.selected_paper

            # Get paper details from backend
            details = rag.get_paper_details(paper_filename)

            if details:
                # TOP BAR: Back button + Open PDF button
                col_back, col_pdf = st.columns([1, 2])
                with col_back:
                    if st.button("← Back to Library", use_container_width=True):
                        print(f"\n[TIMING] Back button clicked - clearing selected_paper", file=sys.stderr, flush=True)
                        st.session_state.selected_paper = None
                        st.rerun()

                with col_pdf:
                    if rag.check_pdf_exists(paper_filename):
                        # Start PDF server (lazy initialization)
                        from lib import pdf_server
                        pdf_server.start_pdf_server()

                        # Get URL for the PDF
                        pdf_url = pdf_server.get_pdf_url(paper_filename)

                        # Create button that opens PDF in new tab
                        st.markdown(f"""
                        <style>
                        .pdf-open-button {{
                            display: inline-block;
                            width: 100%;
                            padding: 0.5rem 0.75rem;
                            background-color: #ff4b4b;
                            color: white;
                            text-align: center;
                            text-decoration: none;
                            border-radius: 0.5rem;
                            font-weight: 600;
                            font-size: 1rem;
                            transition: background-color 0.2s;
                        }}
                        .pdf-open-button:hover {{
                            background-color: #ff6b6b;
                            text-decoration: none;
                        }}
                        </style>
                        <a href="{pdf_url}" target="_blank" class="pdf-open-button">📄 Open PDF</a>
                        """, unsafe_allow_html=True)
                    else:
                        st.button("📄 No PDF Available", use_container_width=True, disabled=True)

                st.divider()

                # TITLE
                display_title = clean_html_from_text(details.get('title', paper_filename.replace('.pdf', '')))
                st.markdown(f"## {display_title}")

                # BIBLIOGRAPHIC INFO SECTION
                st.markdown("### 📚 Bibliographic Information")

                col1, col2 = st.columns(2)

                with col1:
                    # Authors
                    if details.get('authors') and details['authors'][0]:
                        authors_list = [a.strip() for a in details['authors'] if a.strip()]
                        st.markdown("**Authors:**")
                        for author in authors_list[:10]:  # Show up to 10 authors
                            st.markdown(f"- {author}")
                        if len(authors_list) > 10:
                            st.caption(f"... and {len(authors_list) - 10} more")
                    else:
                        st.markdown("**Authors:** _Not available_")

                    # Year
                    if details.get('year'):
                        st.markdown(f"**Year:** {details['year']}")

                    # Paper Type
                    if details.get('paper_type'):
                        st.markdown(f"**Type:** {details['paper_type'].title()}")

                    # Date Added
                    if details.get('date_added'):
                        try:
                            from datetime import datetime
                            date_str = details['date_added']
                            formatted_date = ''
                            # Try multiple formats
                            for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"]:
                                try:
                                    dt = datetime.strptime(date_str.split('.')[0] if '.' in date_str else date_str, fmt)
                                    formatted_date = dt.strftime("%b %d, %Y")
                                    break
                                except:
                                    continue
                            st.markdown(f"**Added:** {formatted_date if formatted_date else date_str}")
                        except:
                            st.markdown(f"**Added:** {details['date_added']}")

                with col2:
                    # Journal
                    if details.get('journal'):
                        st.markdown(f"**Journal:** {details['journal']}")
                    else:
                        st.markdown("**Journal:** _Not available_")

                    # DOI with Find DOI button
                    doi = details.get('doi', '')
                    doi_col1, doi_col2 = st.columns([3, 1])
                    with doi_col1:
                        if doi:
                            st.markdown(f"**DOI:** [{doi}](https://doi.org/{doi})")
                        else:
                            st.markdown("**DOI:** _Not available_")
                    with doi_col2:
                        if not doi and details.get('title'):
                            if st.button("🔍 Find DOI", key=f"find_doi_{paper_filename}", use_container_width=True):
                                # Step 1: Find DOI
                                with st.spinner("Searching Semantic Scholar..."):
                                    from lib.app_helpers import find_doi_via_semantic_scholar, query_crossref_for_metadata
                                    found_doi = find_doi_via_semantic_scholar(details['title'])

                                if found_doi:
                                    # Step 2: Enrich from CrossRef
                                    with st.spinner(f"Enriching metadata from CrossRef..."):
                                        metadata_file = Path("data/metadata.json")
                                        if metadata_file.exists():
                                            with open(metadata_file, 'r', encoding='utf-8') as f:
                                                all_metadata = json.load(f)

                                            if paper_filename in all_metadata:
                                                # Get full metadata from CrossRef
                                                crossref_data = query_crossref_for_metadata(found_doi)

                                                # Update metadata
                                                all_metadata[paper_filename]['doi'] = found_doi

                                                if crossref_data:
                                                    # Extract and save additional metadata
                                                    if crossref_data.get('title'):
                                                        all_metadata[paper_filename]['title'] = crossref_data['title']
                                                    if crossref_data.get('authors'):
                                                        all_metadata[paper_filename]['authors'] = crossref_data['authors']
                                                    if crossref_data.get('year'):
                                                        all_metadata[paper_filename]['year'] = crossref_data['year']
                                                    if crossref_data.get('journal'):
                                                        all_metadata[paper_filename]['journal'] = crossref_data['journal']
                                                    if crossref_data.get('volume'):
                                                        all_metadata[paper_filename]['volume'] = crossref_data['volume']
                                                    if crossref_data.get('issue'):
                                                        all_metadata[paper_filename]['issue'] = crossref_data['issue']
                                                    if crossref_data.get('pages'):
                                                        all_metadata[paper_filename]['pages'] = crossref_data['pages']

                                                    print(f"[FIND DOI] Enriched {paper_filename} with full metadata from CrossRef")

                                                # Save to metadata.json
                                                with open(metadata_file, 'w', encoding='utf-8') as f:
                                                    json.dump(all_metadata, f, indent=2, ensure_ascii=False)
                                                print(f"[FIND DOI] Saved enriched metadata for {paper_filename}")

                                                # Update ChromaDB with all metadata
                                                from lib.rag import DatabaseClient
                                                DatabaseClient.update_paper_metadata(paper_filename, all_metadata[paper_filename])
                                                print(f"[FIND DOI] Updated ChromaDB for {paper_filename}")

                                                # Clear caches
                                                DatabaseClient.clear_cache()
                                                st.cache_data.clear()
                                                st.session_state.reload_papers = True

                                                st.success(f"✅ Found DOI and enriched metadata from CrossRef")
                                                st.rerun()
                                else:
                                    st.warning("❌ No DOI found on Semantic Scholar")

                    # Application
                    if details.get('application'):
                        st.markdown(f"**Application:** {details['application'].title()}")

                    # Reference Count
                    ref_count = len(details.get('references', []))
                    if ref_count > 0:
                        st.markdown(f"**References:** {ref_count}")

                st.divider()

                # TAGS SECTION
                st.markdown("### 🏷️ Tags")

                # Author keywords
                author_keywords = details.get('author_keywords', [])
                if author_keywords and author_keywords[0]:
                    st.markdown("**Author Keywords:**")
                    author_tags_html = []
                    for keyword in author_keywords:
                        if keyword:
                            author_tags_html.append(f'<span class="tag-pill tag-author-keyword">{keyword}</span>')
                    if author_tags_html:
                        st.markdown('<div style="margin: 8px 0;">' + ''.join(author_tags_html) + '</div>', unsafe_allow_html=True)

                # AI-generated tags
                ai_tags_exist = False
                ai_tags_html = []

                # Chemistry tags
                if details.get('chemistries') and details['chemistries'][0]:
                    ai_tags_exist = True
                    for chem in details['chemistries']:
                        if chem:
                            ai_tags_html.append(f'<span class="tag-pill tag-chemistry">{chem}</span>')

                # Topic tags
                if details.get('topics') and details['topics'][0]:
                    ai_tags_exist = True
                    for topic in details['topics']:
                        if topic:
                            ai_tags_html.append(f'<span class="tag-pill tag-topic">{topic}</span>')

                if ai_tags_exist:
                    st.markdown("**AI-Generated Tags:**")
                    st.markdown('<div style="margin: 8px 0;">' + ''.join(ai_tags_html) + '</div>', unsafe_allow_html=True)

                if not author_keywords and not ai_tags_exist:
                    st.caption("_No tags available_")

                st.divider()

                # ABSTRACT SECTION (placeholder for now)
                st.markdown("### 📄 Abstract")
                if details.get('abstract'):
                    st.markdown(details['abstract'])
                else:
                    st.caption("_Abstract not yet extracted. This will be added in a future update._")

                st.divider()

                # NOTES SECTION (editable)
                st.markdown("### 📝 Notes")

                # Load notes from a notes file (or session state)
                notes_file = Path(f"data/notes/{paper_filename}.txt")
                notes_file.parent.mkdir(parents=True, exist_ok=True)

                current_notes = ""
                if notes_file.exists():
                    with open(notes_file, 'r', encoding='utf-8') as f:
                        current_notes = f.read()

                notes = st.text_area(
                    "Your notes about this paper:",
                    value=current_notes,
                    height=200,
                    key=f"notes_{paper_filename}",
                    placeholder="Add your notes, thoughts, or important findings here..."
                )

                col_save, col_clear = st.columns([1, 4])
                with col_save:
                    if st.button("💾 Save Notes", use_container_width=True):
                        with open(notes_file, 'w', encoding='utf-8') as f:
                            f.write(notes)
                        st.toast("Notes saved!", icon="✅")

                st.divider()

                # COLLECTIONS SECTION
                st.markdown("### 📁 Collections")

                # Get current collections for this paper
                current_collections = collections.get_paper_collections(paper_filename)
                all_collections_list = collections.get_all_collections()

                # Display current collections as color-coded tags
                if current_collections:
                    cols_display = st.columns(min(len(current_collections), 4))
                    for idx, coll in enumerate(current_collections):
                        with cols_display[idx % 4]:
                            color = coll.get('color') or '#6c757d'
                            st.markdown(
                                f'<span style="display: inline-block; background-color: {color}; color: white; '
                                f'padding: 4px 12px; border-radius: 12px; font-size: 13px; margin: 2px;">'
                                f'{coll["name"]}</span>',
                                unsafe_allow_html=True
                            )
                else:
                    st.caption("Not in any collections")

                # Add/Remove from collection controls
                col_add, col_remove = st.columns(2)

                with col_add:
                    st.markdown("**Add to Collection:**")
                    # Filter out collections the paper is already in
                    current_collection_ids = {c['id'] for c in current_collections}
                    available_collections = [c for c in all_collections_list if c['id'] not in current_collection_ids]

                    if available_collections:
                        add_col1, add_col2 = st.columns([3, 1])
                        with add_col1:
                            selected_to_add = st.selectbox(
                                "Select collection",
                                options=[c['name'] for c in available_collections],
                                key=f"add_collection_{paper_filename}",
                                label_visibility="collapsed"
                            )
                        with add_col2:
                            if st.button("➕", key=f"btn_add_{paper_filename}", use_container_width=True, help="Add to collection"):
                                collection_to_add = next((c for c in available_collections if c['name'] == selected_to_add), None)
                                if collection_to_add:
                                    result = collections.add_paper_to_collection(collection_to_add['id'], paper_filename)
                                    if result['success']:
                                        st.toast(f"Added to '{collection_to_add['name']}'", icon="✅")
                                        st.rerun()
                                    else:
                                        st.error(result['message'])
                    else:
                        st.caption("All collections added" if all_collections_list else "No collections available")

                with col_remove:
                    st.markdown("**Remove from Collection:**")
                    if current_collections:
                        remove_col1, remove_col2 = st.columns([3, 1])
                        with remove_col1:
                            selected_to_remove = st.selectbox(
                                "Select collection",
                                options=[c['name'] for c in current_collections],
                                key=f"remove_collection_{paper_filename}",
                                label_visibility="collapsed"
                            )
                        with remove_col2:
                            if st.button("➖", key=f"btn_remove_{paper_filename}", use_container_width=True, help="Remove from collection"):
                                collection_to_remove = next((c for c in current_collections if c['name'] == selected_to_remove), None)
                                if collection_to_remove:
                                    result = collections.remove_paper_from_collection(collection_to_remove['id'], paper_filename)
                                    if result['success']:
                                        st.toast(f"Removed from '{collection_to_remove['name']}'", icon="✅")
                                        st.rerun()
                                    else:
                                        st.error(result['message'])
                    else:
                        st.caption("Not in any collections")

                # Create new collection expander
                with st.expander("➕ Create New Collection"):
                    new_coll_name = st.text_input(
                        "Collection Name",
                        placeholder="e.g., SOH Methods, Grant Proposal, EIS Papers",
                        key=f"new_coll_name_{paper_filename}"
                    )
                    new_coll_color = st.color_picker(
                        "Collection Color",
                        value="#6c757d",
                        key=f"new_coll_color_{paper_filename}"
                    )
                    new_coll_desc = st.text_area(
                        "Description (optional)",
                        placeholder="Brief description of this collection...",
                        height=80,
                        key=f"new_coll_desc_{paper_filename}"
                    )

                    if st.button("Create Collection", key=f"create_coll_{paper_filename}", type="primary"):
                        if new_coll_name.strip():
                            result = collections.create_collection(
                                new_coll_name.strip(),
                                new_coll_color,
                                new_coll_desc.strip()
                            )
                            if result['success']:
                                # Automatically add current paper to the new collection
                                collections.add_paper_to_collection(result['id'], paper_filename)
                                st.toast(f"Collection '{new_coll_name}' created and paper added!", icon="✅")
                                st.rerun()
                            else:
                                st.error(result['message'])
                        else:
                            st.warning("Please enter a collection name")

                st.divider()

                # REFERENCES SECTION
                references = details.get('references', [])
                if references:
                    with st.expander(f"📚 References ({len(references)})", expanded=False):
                        st.caption("Papers cited by this work")

                        # Get all DOIs in the library for status checking
                        # Normalize DOIs by removing URL prefix if present
                        def normalize_doi(doi_string):
                            """Remove https://doi.org/ prefix and lowercase"""
                            if not doi_string:
                                return ''
                            doi = doi_string.lower().strip()
                            # Remove common DOI URL prefixes
                            for prefix in ['https://doi.org/', 'http://doi.org/', 'doi.org/', 'doi:']:
                                if doi.startswith(prefix):
                                    doi = doi[len(prefix):]
                            return doi

                        library_dois = {normalize_doi(p.get('doi', '')) for p in papers if p.get('doi')}
                        library_dois.discard('')  # Remove empty strings

                        # Filter and prepare references data
                        refs_data = []
                        refs_full_data = {}  # Store full ref data separately (not in DataFrame)
                        for i, ref in enumerate(references):
                            # Get reference fields
                            title = ref.get('article-title', '').strip()
                            authors = ref.get('author', '').strip()

                            # Mark incomplete references (missing title or author)
                            is_incomplete = not title or not authors

                            # Show what data we have, even if incomplete
                            display_title = title if title else '(No title)'
                            display_authors = authors if authors else '(No author)'

                            # Check if in library (only for complete references)
                            in_library = False
                            if not is_incomplete:
                                # Method 1: Check by normalized DOI
                                ref_doi = normalize_doi(ref.get('DOI', ''))
                                if ref_doi and ref_doi in library_dois:
                                    in_library = True
                                # Method 2: If no DOI, check by title match (fuzzy)
                                elif title:
                                    from difflib import SequenceMatcher
                                    title_lower = title.lower()
                                    for p in papers:
                                        p_title = p.get('title', '').lower()
                                        if p_title and SequenceMatcher(None, title_lower, p_title).ratio() > 0.9:
                                            in_library = True
                                            break

                            # Format journal with volume
                            journal = ref.get('journal-title', '')
                            if journal and ref.get('volume'):
                                journal = f"{journal}, Vol. {ref['volume']}"

                            row_idx = len(refs_data)
                            refs_data.append({
                                'Title': display_title,
                                'Authors': display_authors,
                                'Year': str(ref.get('year', '')) if ref.get('year') else '—',
                                'Journal': journal if journal else '—',
                                'DOI': ref.get('DOI', '—'),
                                'Status': '✓ In Library' if in_library else ('Incomplete' if is_incomplete else 'Not in Library'),
                                '_in_library': in_library,
                                '_incomplete': is_incomplete
                            })
                            # Store full ref data separately (dicts break AgGrid rendering)
                            refs_full_data[row_idx] = ref

                        if not refs_data:
                            st.info("No valid references found (references must have both title and author)")
                        else:
                            # Build DataFrame
                            refs_df = pd.DataFrame(refs_data)

                            # Configure AG Grid (EXACT copy from library table)
                            refs_gb = GridOptionsBuilder.from_dataframe(refs_df)

                            # Title column - EXACT copy from library table
                            refs_gb.configure_column("Title",
                                flex=3,
                                minWidth=250,
                                wrapText=True,
                                autoHeight=False,
                                resizable=True,
                                cellStyle={
                                    'whiteSpace': 'normal !important',
                                    'lineHeight': '1.4 !important',
                                    'display': '-webkit-box !important',
                                    '-webkit-line-clamp': '2 !important',
                                    '-webkit-box-orient': 'vertical !important',
                                    'overflow': 'hidden !important',
                                    'textOverflow': 'ellipsis !important',
                                    'padding': '8px !important',
                                    'maxHeight': '45px !important'
                                },
                                tooltipField="Title"
                            )

                            # Authors column - EXACT copy from library table
                            refs_gb.configure_column("Authors",
                                flex=2,
                                minWidth=180,
                                wrapText=True,
                                autoHeight=False,
                                resizable=True,
                                cellStyle={
                                    'whiteSpace': 'normal !important',
                                    'lineHeight': '1.4 !important',
                                    'display': '-webkit-box !important',
                                    '-webkit-line-clamp': '2 !important',
                                    '-webkit-box-orient': 'vertical !important',
                                    'overflow': 'hidden !important',
                                    'textOverflow': 'ellipsis !important',
                                    'padding': '8px !important',
                                    'maxHeight': '45px !important'
                                },
                                tooltipField="Authors"
                            )

                            # Year column - EXACT copy from library table
                            refs_gb.configure_column("Year",
                                width=70,
                                minWidth=60,
                                maxWidth=90,
                                resizable=True,
                                cellStyle={
                                    'whiteSpace': 'nowrap !important',
                                    'overflow': 'hidden !important',
                                    'textOverflow': 'ellipsis !important',
                                    'display': 'flex !important',
                                    'justifyContent': 'flex-start !important',
                                    'alignItems': 'center !important',
                                    'textAlign': 'left !important',
                                    'paddingLeft': '8px !important'
                                }
                            )

                            # Journal column - EXACT copy from library table
                            refs_gb.configure_column("Journal",
                                flex=2,
                                minWidth=150,
                                wrapText=True,
                                resizable=True,
                                cellStyle={
                                    'whiteSpace': 'normal !important',
                                    'lineHeight': '1.4 !important',
                                    'display': '-webkit-box !important',
                                    '-webkit-line-clamp': '2 !important',
                                    '-webkit-box-orient': 'vertical !important',
                                    'overflow': 'hidden !important',
                                    'textOverflow': 'ellipsis !important',
                                    'padding': '8px !important',
                                    'maxHeight': '45px !important'
                                },
                                tooltipField="Journal"
                            )

                            # DOI column - EXACT copy from library table
                            refs_gb.configure_column("DOI",
                                flex=1.5,
                                minWidth=140,
                                resizable=True,
                                cellStyle={
                                    'overflow': 'hidden'
                                },
                                tooltipField="DOI"
                            )

                            # Status column
                            refs_gb.configure_column("Status",
                                width=130,
                                minWidth=110,
                                maxWidth=150,
                                resizable=False,
                                cellStyle={'textAlign': 'center', 'overflow': 'hidden'}
                            )

                            # Hide internal columns
                            refs_gb.configure_column("_in_library", hide=True)

                            # Grid options - EXACT copy from library table
                            refs_gb.configure_selection(selection_mode='single', use_checkbox=False)
                            refs_gb.configure_grid_options(
                                headerHeight=40,
                                suppressRowHoverHighlight=False,
                                enableCellTextSelection=True,
                                ensureDomOrder=True,
                                domLayout='normal',
                                rowHeight=60,
                                suppressHorizontalScroll=True,
                                suppressColumnVirtualisation=False,
                                suppressRowVirtualisation=False
                            )

                            refs_grid_options = refs_gb.build()

                            # Custom CSS - EXACT copy from library table
                            if st.session_state.theme == 'dark':
                                refs_custom_css = {
                                    ".ag-header-cell-label": {
                                        "font-weight": "600",
                                        "font-size": "14px",
                                        "font-family": "inherit",
                                        "color": "#FFFFFF !important"
                                    },
                                    ".ag-header-cell-text": {
                                        "color": "#FFFFFF !important",
                                        "font-size": "14px"
                                    },
                                    ".ag-header": {
                                        "background-color": "#262730 !important",
                                        "border-bottom": "1px solid #444444 !important",
                                        "height": "40px !important"
                                    },
                                    ".ag-header-cell": {
                                        "background-color": "#262730 !important",
                                        "color": "#FFFFFF !important",
                                        "padding": "0 8px !important"
                                    },
                                    ".ag-root-wrapper": {
                                        "background-color": "#1E1E1E !important",
                                        "width": "100% !important",
                                        "font-family": "inherit",
                                        "font-size": "14px"
                                    },
                                    ".ag-row": {
                                        "border-bottom": "1px solid #444444 !important",
                                        "background-color": "#1E1E1E !important",
                                        "height": "60px !important"
                                    },
                                    ".ag-cell": {
                                        "padding": "8px !important",
                                        "display": "flex !important",
                                        "align-items": "center !important",
                                        "color": "#E0E0E0 !important",
                                        "overflow": "hidden !important",
                                        "text-overflow": "ellipsis !important",
                                        "font-size": "14px !important",
                                        "font-family": "inherit !important",
                                        "line-height": "1.5 !important"
                                    },
                                    ".ag-cell .ag-cell-value": {
                                        "display": "-webkit-box !important",
                                        "-webkit-line-clamp": "2 !important",
                                        "-webkit-box-orient": "vertical !important",
                                        "overflow": "hidden !important",
                                        "text-overflow": "ellipsis !important",
                                        "white-space": "normal !important",
                                        "line-height": "1.4 !important",
                                        "max-height": "42px !important",
                                        "width": "100% !important"
                                    },
                                    ".ag-row-hover": {"background-color": "#2D2D2D !important"},
                                    ".ag-center-cols-viewport": {"background-color": "#1E1E1E !important"},
                                    ".ag-body-viewport": {"background-color": "#1E1E1E !important"},
                                }
                            else:
                                refs_custom_css = {
                                    ".ag-header-cell-label": {
                                        "font-weight": "600",
                                        "font-size": "14px",
                                        "font-family": "inherit",
                                        "color": "#2c3e50 !important"
                                    },
                                    ".ag-header-cell-text": {
                                        "color": "#2c3e50 !important",
                                        "font-size": "14px"
                                    },
                                    ".ag-header": {
                                        "background-color": "#F0F2F6 !important",
                                        "border-bottom": "1px solid #D0D0D0 !important",
                                        "height": "40px !important"
                                    },
                                    ".ag-header-cell": {
                                        "background-color": "#F0F2F6 !important",
                                        "color": "#2c3e50 !important",
                                        "padding": "0 8px !important"
                                    },
                                    ".ag-root-wrapper": {
                                        "background-color": "#FFFFFF !important",
                                        "width": "100% !important",
                                        "font-family": "inherit",
                                        "font-size": "14px"
                                    },
                                    ".ag-row": {
                                        "border-bottom": "1px solid #ecf0f1 !important",
                                        "background-color": "#FFFFFF !important",
                                        "height": "60px !important"
                                    },
                                    ".ag-cell": {
                                        "padding": "8px !important",
                                        "display": "flex !important",
                                        "align-items": "center !important",
                                        "color": "#262730 !important",
                                        "overflow": "hidden !important",
                                        "text-overflow": "ellipsis !important",
                                        "font-size": "14px !important",
                                        "font-family": "inherit !important",
                                        "line-height": "1.5 !important"
                                    },
                                    ".ag-cell .ag-cell-value": {
                                        "display": "-webkit-box !important",
                                        "-webkit-line-clamp": "2 !important",
                                        "-webkit-box-orient": "vertical !important",
                                        "overflow": "hidden !important",
                                        "text-overflow": "ellipsis !important",
                                        "white-space": "normal !important",
                                        "line-height": "1.4 !important",
                                        "max-height": "42px !important",
                                        "width": "100% !important"
                                    },
                                    ".ag-row-hover": {"background-color": "#f8f9fa !important"},
                                    ".ag-center-cols-viewport": {"background-color": "#FFFFFF !important"},
                                    ".ag-body-viewport": {"background-color": "#FFFFFF !important"},
                                }

                            st.divider()

                            # Prepare data for buttons (need counts before displaying grid)
                            missing_refs = refs_df[(refs_df['_in_library'] == False) & (refs_df['_incomplete'] == False)]
                            incomplete_count = refs_df[refs_df['_incomplete'] == True].shape[0]

                            # Display references table with 1-based indexing
                            refs_display = refs_df.drop(columns=['_in_library', '_incomplete']).copy()
                            refs_display.insert(0, '#', range(1, len(refs_display) + 1))

                            # Minimal AgGrid with ONLY 2-line clamp CSS added
                            refs_minimal_gb = GridOptionsBuilder.from_dataframe(refs_display)

                            # Configure # column to be narrow and centered
                            refs_minimal_gb.configure_column("#",
                                width=60,
                                maxWidth=80,
                                pinned='left',
                                cellStyle={'textAlign': 'center', 'fontWeight': '600'}
                            )

                            # Apply conditional styling based on Status (incomplete references)
                            incomplete_style = JsCode("""
                                function(params) {
                                    if (params.data.Status === 'Incomplete') {
                                        return {
                                            'color': '#999999',
                                            'fontStyle': 'italic'
                                        };
                                    }
                                    return {};
                                }
                            """)

                            # Apply greyed-out styling to all columns for incomplete rows
                            for col in ['#', 'Title', 'Authors', 'Year', 'Journal', 'DOI', 'Status']:
                                refs_minimal_gb.configure_column(col, cellStyle=incomplete_style)

                            # Enable multi-selection with checkboxes
                            refs_minimal_gb.configure_selection(selection_mode='multiple', use_checkbox=True)
                            refs_minimal_gb.configure_grid_options(domLayout='autoHeight')
                            refs_minimal_options = refs_minimal_gb.build()

                            # ONLY add the critical CSS for 2-line text wrapping (from library table)
                            refs_minimal_css = {
                                ".ag-cell .ag-cell-value": {
                                    "display": "-webkit-box !important",
                                    "-webkit-line-clamp": "2 !important",
                                    "-webkit-box-orient": "vertical !important",
                                    "overflow": "hidden !important",
                                    "text-overflow": "ellipsis !important",
                                    "white-space": "normal !important",
                                    "line-height": "1.4 !important",
                                    "max-height": "42px !important"
                                }
                            }

                            # Add to Library buttons - BEFORE grid for better UX
                            # Use container to render buttons at top but access grid response
                            buttons_placeholder = st.container()

                            refs_grid_response = AgGrid(
                                refs_display,
                                gridOptions=refs_minimal_options,
                                custom_css=refs_minimal_css,
                                fit_columns_on_grid_load=True,
                                theme='streamlit',
                                allow_unsafe_jscode=True,
                                key=f"refs_minimal_{paper_filename}"
                            )

                            # Render buttons in the placeholder at the top
                            with buttons_placeholder:
                                if len(missing_refs) > 0 or (refs_grid_response.get('selected_rows') is not None and len(refs_grid_response.get('selected_rows', [])) > 0):
                                    col1, col2 = st.columns(2)

                                    # "Add Selected" button - primary action
                                    with col1:
                                        selected_rows = refs_grid_response.get('selected_rows', [])
                                        if selected_rows is not None and len(selected_rows) > 0:
                                            # Filter selected rows to only include those not in library and not incomplete
                                            selected_df = pd.DataFrame(selected_rows)
                                            addable_selected = []
                                            for _, row in selected_df.iterrows():
                                                # Find the original row index based on # column
                                                ref_num = row['#'] - 1  # Convert back to 0-based index
                                                if ref_num in refs_full_data:
                                                    # Check if this ref is not in library and not incomplete
                                                    orig_row = refs_df.iloc[ref_num]
                                                    if not orig_row['_in_library'] and not orig_row['_incomplete']:
                                                        addable_selected.append(ref_num)

                                            if len(addable_selected) > 0:
                                                if st.button(f"➕ Add Selected ({len(addable_selected)})", type="primary", use_container_width=True, key=f"add_selected_{paper_filename}"):
                                                    progress_bar = st.progress(0)
                                                    success_count = 0

                                                    for idx, ref_idx in enumerate(addable_selected):
                                                        ref_data = refs_full_data[ref_idx]
                                                        result = import_reference(ref_data)
                                                        if result['success']:
                                                            success_count += 1
                                                        progress_bar.progress((idx + 1) / len(addable_selected))

                                                    st.toast(f"✅ Added {success_count} of {len(addable_selected)} references", icon="✅")
                                                    st.rerun()
                                            else:
                                                st.info("Selected references are already in library or incomplete")
                                        else:
                                            st.caption("Select references using checkboxes to add them")

                                    # "Add All Missing" button - secondary action
                                    with col2:
                                        if len(missing_refs) > 0:
                                            if incomplete_count > 0:
                                                st.caption(f"{len(missing_refs)} missing ({incomplete_count} incomplete excluded)")
                                            if st.button(f"➕ Add All Missing ({len(missing_refs)})", type="secondary", use_container_width=True, key=f"add_all_missing_{paper_filename}"):
                                                progress_bar = st.progress(0)
                                                success_count = 0

                                                for idx, (row_idx, row) in enumerate(missing_refs.iterrows()):
                                                    if row_idx not in refs_full_data:
                                                        continue

                                                    ref_data = refs_full_data[row_idx]
                                                    result = import_reference(ref_data)
                                                    if result['success']:
                                                        success_count += 1
                                                    progress_bar.progress((idx + 1) / len(missing_refs))

                                                st.toast(f"✅ Added {success_count} of {len(missing_refs)} references", icon="✅")
                                                st.rerun()
                                        else:
                                            st.caption("All references already in library")

                                    st.divider()

                # PDF UPLOAD (if no PDF exists)
                if not rag.check_pdf_exists(paper_filename):
                    with st.expander("📤 Upload PDF", expanded=True):
                        st.info("No PDF file found for this paper. Upload one to enable PDF viewing.")

                        uploaded_pdf = st.file_uploader(
                            "Upload PDF",
                            type=['pdf'],
                            key=f"upload_pdf_{paper_filename}"
                        )

                        if uploaded_pdf:
                            # Save uploaded PDF
                            papers_dir = Path("papers")
                            papers_dir.mkdir(parents=True, exist_ok=True)
                            pdf_path = papers_dir / paper_filename

                            with open(pdf_path, 'wb') as f:
                                f.write(uploaded_pdf.read())

                            st.success("✅ PDF uploaded successfully!")
                            st.info("Processing PDF... This may take a moment.")

                            # TODO: Run ingestion pipeline on the uploaded PDF
                            time.sleep(1)
                            st.rerun()

                st.divider()

                # EDIT METADATA SECTION (at the bottom)
                with st.expander("Edit Metadata", expanded=False):
                    current_doi = details.get('doi', '')
                    new_doi = st.text_input(
                        "DOI",
                        value=current_doi,
                        placeholder="10.xxxx/xxxxx",
                        help="Enter or update the DOI for this paper",
                        key=f"doi_edit_{paper_filename}"
                    )

                    col_btn1, col_btn2 = st.columns(2)

                    with col_btn1:
                        if st.button("🔄 Refresh from CrossRef", key=f"refresh_{paper_filename}", use_container_width=True):
                            doi_to_use = new_doi.strip() if new_doi.strip() else current_doi
                            if doi_to_use:
                                with st.spinner(f'Querying CrossRef for DOI: {doi_to_use}...'):
                                    crossref_metadata = query_crossref_for_metadata(doi_to_use)

                                if crossref_metadata:
                                    success = update_paper_metadata(paper_filename, doi_to_use, crossref_metadata)
                                    if success:
                                        st.toast('✅ Metadata updated from CrossRef!', icon='✅')
                                        st.rerun()
                                else:
                                    st.warning("No metadata found in CrossRef for this DOI")
                            else:
                                st.warning("Please enter a DOI first")

                    with col_btn2:
                        if st.button("💾 Save", key=f"save_doi_{paper_filename}", use_container_width=True):
                            if new_doi.strip() and new_doi.strip() != current_doi:
                                metadata_file = Path("data/metadata.json")
                                if metadata_file.exists():
                                    with open(metadata_file, 'r', encoding='utf-8') as f:
                                        all_metadata = json.load(f)

                                    if paper_filename in all_metadata:
                                        all_metadata[paper_filename]['doi'] = new_doi.strip()

                                        with open(metadata_file, 'w', encoding='utf-8') as f:
                                            json.dump(all_metadata, f, indent=2, ensure_ascii=False)

                                        # Update ChromaDB
                                        update_chromadb_metadata(paper_filename, all_metadata[paper_filename])

                                        st.toast('✅ DOI saved!', icon='✅')
                                        st.rerun()
                            elif new_doi.strip() == current_doi:
                                st.info("DOI unchanged")
                            else:
                                st.warning("Please enter a DOI")

        else:
            # Table view
            st.subheader("Paper Library")

            # Search box and enrichment button
            col_search, col_enrich = st.columns([3, 1])
            with col_search:
                search_query = st.text_input(
                    "Search papers",
                    placeholder="Search by title, authors, journal...",
                    label_visibility="collapsed",
                    key="library_search"
                )
            with col_enrich:
                if st.button("🔍 Enrich Metadata", help="Fetch missing metadata from CrossRef for papers with URLs", use_container_width=True):
                    # Progress containers
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    def progress_update(current, total, title):
                        progress = current / total if total > 0 else 0
                        progress_bar.progress(progress)
                        status_text.text(f"Enriching paper {current} of {total}: {title[:50]}...")

                    result = enrich_library_metadata(max_papers=None, progress_callback=progress_update)

                    # Clear progress indicators
                    progress_bar.empty()
                    status_text.empty()

                    if result['success']:
                        # Summary message
                        enriched = result.get('enriched', 0)
                        failed = result.get('failed', 0)
                        total = result.get('total', 0)
                        skipped = total - enriched - failed

                        summary_parts = []
                        if enriched > 0:
                            summary_parts.append(f"✅ Enriched {enriched} paper{'s' if enriched != 1 else ''}")
                            st.session_state.reload_papers = True  # Invalidate cache after enrichment
                        if failed > 0:
                            summary_parts.append(f"⚠️ {failed} failed")
                        if skipped > 0:
                            summary_parts.append(f"⏭️ {skipped} skipped")

                        if enriched > 0:
                            st.success(". ".join(summary_parts) + ".")
                        elif total > 0:
                            st.info(". ".join(summary_parts) + ".")
                        else:
                            st.info("No papers need enrichment")

                        # Details in expander
                        if result.get('logs'):
                            with st.expander("Show details"):
                                for log in result['logs']:
                                    st.text(log)

                        if enriched > 0:
                            time.sleep(2)
                            st.rerun()
                    else:
                        st.error(result.get('message', 'Enrichment failed'))

            # Horizontal filter bar
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                filter_chemistry = st.selectbox(
                    "Chemistry",
                    options=["All Chemistries"] + filter_options['chemistries'],
                    key="library_filter_chemistry"
                )
            with col2:
                filter_topic = st.selectbox(
                    "Topic",
                    options=["All Topics"] + filter_options['topics'],
                    key="library_filter_topic"
                )
            with col3:
                filter_paper_type = st.selectbox(
                    "Paper Type",
                    options=["All Types"] + filter_options['paper_types'],
                    key="library_filter_paper_type"
                )
            with col4:
                all_collections = collections.get_all_collections()
                collection_options = ["All Collections"] + [c['name'] for c in all_collections]
                filter_collection = st.selectbox(
                    "📁 Collection",
                    options=collection_options,
                    key="library_filter_collection"
                )
            with col5:
                filter_status = st.selectbox(
                    "Status",
                    options=["All Papers", "✅ Complete", "📋 Metadata Only", "⚠️ Incomplete"],
                    key="library_filter_status"
                )

            # Build library DataFrame using cached function (filters + formats data)
            _df_start = timing_module.time()
            df = cached_operations.build_library_dataframe(
                papers=papers,
                search_query=search_query or "",
                filter_chemistry=filter_chemistry or "All Chemistries",
                filter_topic=filter_topic or "All Topics",
                filter_paper_type=filter_paper_type or "All Types",
                filter_collection=filter_collection or "All Collections",
                filter_status=filter_status or "All Papers"
            )
            print(f"[TIMING] DataFrame built: {timing_module.time() - _df_start:.3f}s (total: {timing_module.time() - _start_time:.3f}s)", file=sys.stderr, flush=True)

            # Count filtered papers
            filtered_count = len(df)
            st.write(f"Showing {filtered_count} of {len(papers)} papers")

            # Action buttons
            if len(df) > 0:
                st.caption("💡 **Tip:** Click a row to view details • Use checkboxes for bulk delete")
                btn_col1, spacer_col = st.columns([1, 4])
                with btn_col1:
                    delete_button = st.button("🗑️ Delete Selected", type="secondary", use_container_width=True)

            # Configure AG Grid with flex sizing for full-width layout
            gb = GridOptionsBuilder.from_dataframe(df)

            # Configure column properties with flex sizing to fill container width
            # Status column with emoji indicators (colorblind-friendly)
            status_cell_renderer = JsCode("""
                class StatusCellRenderer {
                    init(params) {
                        this.eGui = document.createElement('div');
                        this.eGui.style.textAlign = 'center';
                        this.eGui.style.padding = '8px';
                        this.eGui.style.fontSize = '14px';
                        this.eGui.textContent = params.value;
                    }

                    getGui() {
                        return this.eGui;
                    }
                }
            """)

            # Configure Select (checkbox) column
            gb.configure_column("Select",
                headerName="",
                width=50,
                minWidth=50,
                maxWidth=50,
                checkboxSelection=True,
                headerCheckboxSelection=True,
                resizable=False,
                suppressMenu=True,
                lockPosition=True,
                pinned='left',
                valueFormatter=JsCode("function(params) { return ''; }"),  # Don't display value, just checkbox
                cellStyle={
                    'display': 'flex',
                    'alignItems': 'center',
                    'justifyContent': 'center'
                }
            )

            # Custom comparator for Status column (Complete → Metadata Only → Incomplete)
            status_comparator = JsCode("""
                function(valueA, valueB) {
                    const order = {
                        '✅ Complete': 1,
                        '📋 Metadata Only': 2,
                        '⚠️ Incomplete': 3
                    };
                    return (order[valueA] || 999) - (order[valueB] || 999);
                }
            """)

            gb.configure_column("Status",
                width=140,
                minWidth=120,
                maxWidth=160,
                resizable=True,
                cellRenderer=status_cell_renderer,
                comparator=status_comparator,
                cellStyle={
                    'textAlign': 'center',
                    'padding': '4px 8px',
                    'fontSize': '13px',
                    'whiteSpace': 'nowrap',
                    'overflow': 'visible',
                    'display': 'flex',
                    'alignItems': 'center',
                    'justifyContent': 'center'
                }
            )

            gb.configure_column("Title",
                flex=3,  # Takes 3 parts of available space
                minWidth=250,
                wrapText=True,
                autoHeight=False,
                resizable=True,
                cellStyle={
                    'whiteSpace': 'normal !important',
                    'lineHeight': '1.4 !important',
                    'display': '-webkit-box !important',
                    '-webkit-line-clamp': '2 !important',
                    '-webkit-box-orient': 'vertical !important',
                    'overflow': 'hidden !important',
                    'textOverflow': 'ellipsis !important',
                    'padding': '8px !important',
                    'maxHeight': '45px !important'
                },
                tooltipField="Title"
            )
            gb.configure_column("Authors",
                flex=2,  # Takes 2 parts of available space
                minWidth=180,
                wrapText=True,
                autoHeight=False,
                resizable=True,
                cellStyle={
                    'whiteSpace': 'normal !important',
                    'lineHeight': '1.4 !important',
                    'display': '-webkit-box !important',
                    '-webkit-line-clamp': '2 !important',
                    '-webkit-box-orient': 'vertical !important',
                    'overflow': 'hidden !important',
                    'textOverflow': 'ellipsis !important',
                    'padding': '8px !important',
                    'maxHeight': '45px !important'
                },
                tooltipField="Authors"
            )
            gb.configure_column("Year",
                width=70,  # Fixed width for narrow columns
                minWidth=60,
                maxWidth=90,
                resizable=True,
                cellStyle={
                    'whiteSpace': 'nowrap !important',
                    'overflow': 'hidden !important',
                    'textOverflow': 'ellipsis !important',
                    'display': 'flex !important',
                    'justifyContent': 'flex-start !important',
                    'alignItems': 'center !important',
                    'textAlign': 'left !important',
                    'paddingLeft': '8px !important'
                }
            )
            gb.configure_column("Journal",
                flex=2,  # Takes 2 parts of available space
                minWidth=150,
                wrapText=True,
                resizable=True,
                cellStyle={
                    'whiteSpace': 'normal !important',
                    'lineHeight': '1.4 !important',
                    'display': '-webkit-box !important',
                    '-webkit-line-clamp': '2 !important',
                    '-webkit-box-orient': 'vertical !important',
                    'overflow': 'hidden !important',
                    'textOverflow': 'ellipsis !important',
                    'padding': '8px !important',
                    'maxHeight': '45px !important'
                },
                tooltipField="Journal"
            )
            gb.configure_column("Collections",
                headerName='📁 Collections',
                flex=1.5,  # Takes 1.5 parts of available space
                minWidth=120,
                wrapText=True,
                resizable=True,
                cellStyle={
                    'whiteSpace': 'normal !important',
                    'lineHeight': '1.4 !important',
                    'display': '-webkit-box !important',
                    '-webkit-line-clamp': '2 !important',
                    '-webkit-box-orient': 'vertical !important',
                    'overflow': 'hidden !important',
                    'textOverflow': 'ellipsis !important',
                    'padding': '8px !important',
                    'maxHeight': '45px !important',
                    'fontSize': '13px !important'
                },
                tooltipField="Collections"
            )
            gb.configure_column("Added",
                width=110,
                minWidth=100,
                maxWidth=130,
                resizable=True,
                cellStyle={
                    'whiteSpace': 'nowrap !important',
                    'overflow': 'hidden !important',
                    'textOverflow': 'ellipsis !important',
                    'display': 'flex !important',
                    'justifyContent': 'flex-start !important',
                    'alignItems': 'center !important',
                    'textAlign': 'left !important',
                    'paddingLeft': '8px !important',
                    'fontSize': '13px !important'
                }
            )

            # DOI column with clickable link
            doi_renderer = JsCode("""
                class DoiRenderer {
                    init(params) {
                        this.eGui = document.createElement('div');

                        if (!params.value || params.value === '—' || params.value === '') {
                            this.eGui.innerHTML = '<span style="color: #999; font-style: italic;">No DOI</span>';
                        } else {
                            // Value is already just the DOI (10.xxxx/...), URL is in _doi_url
                            const url = params.data._doi_url || 'https://doi.org/' + params.value;
                            this.eGui.innerHTML = '<a href="' + url + '" target="_blank" rel="noopener noreferrer" style="color: #1f77b4; text-decoration: underline;">' + params.value + '</a>';
                        }
                    }

                    getGui() {
                        return this.eGui;
                    }
                }
            """)
            gb.configure_column("DOI",
                flex=1.5,
                minWidth=140,
                resizable=True,
                cellRenderer=doi_renderer,
                cellStyle={
                    'overflow': 'hidden'
                },
                tooltipField="DOI"
            )

            # Read column with actual checkbox
            checkbox_renderer = JsCode("""
                class CheckboxRenderer {
                    init(params) {
                        this.params = params;
                        this.eGui = document.createElement('div');
                        this.eGui.style.textAlign = 'center';
                        this.eGui.style.display = 'flex';
                        this.eGui.style.alignItems = 'center';
                        this.eGui.style.justifyContent = 'center';
                        this.eGui.style.height = '100%';

                        const checkbox = document.createElement('input');
                        checkbox.type = 'checkbox';
                        checkbox.checked = params.value;
                        checkbox.style.cursor = 'pointer';
                        checkbox.style.width = '16px';
                        checkbox.style.height = '16px';

                        // Handle checkbox change
                        checkbox.addEventListener('change', (e) => {
                            e.stopPropagation();  // Don't trigger row navigation
                            params.setValue(e.target.checked);
                        });

                        this.eGui.appendChild(checkbox);
                    }

                    getGui() {
                        return this.eGui;
                    }
                }
            """)
            gb.configure_column("Read",
                width=70,  # Fixed width for checkbox column
                minWidth=60,
                maxWidth=80,
                resizable=False,
                cellRenderer=checkbox_renderer,
                editable=True,
                cellEditor='agCheckboxCellEditor',
                cellStyle={
                    'overflow': 'hidden',
                    'textAlign': 'center'
                }
            )

            # Hide internal columns
            gb.configure_column("_filename", hide=True)
            gb.configure_column("_doi_url", hide=True)
            gb.configure_column("_paper_title", hide=True)

            # Grid options - configured for full-width with virtualization for large datasets
            # Multi-select enabled via Select column (configured above with checkboxSelection=True)
            gb.configure_selection(selection_mode='multiple', suppressRowClickSelection=True)
            gb.configure_grid_options(
                headerHeight=40,
                suppressRowHoverHighlight=False,
                enableCellTextSelection=True,
                ensureDomOrder=True,
                domLayout='normal',  # Fixed height with internal scrolling and virtualization
                rowHeight=60,  # Fixed height to accommodate 2-line titles
                suppressHorizontalScroll=True,  # Disable horizontal scrolling - table fills width
                suppressColumnVirtualisation=False,  # Enable column virtualization
                suppressRowVirtualisation=False,  # Enable row virtualization for performance with many rows
                rowSelection='multiple',  # Enable multi-row selection
                rowMultiSelectWithClick=False,  # Prevent accidental multi-select on row click
            )

            grid_options = gb.build()

            # Add cell click handler to separate different click behaviors
            grid_options['onCellClicked'] = JsCode("""
                function(event) {
                    const colId = event.column ? event.column.colId : null;

                    // 1. SELECT CHECKBOX: Do nothing, let default checkbox behavior handle it
                    if (colId === 'Select') {
                        return;
                    }

                    // 2. DOI COLUMN: Do nothing here, link click is handled by <a> tag
                    if (colId === 'DOI') {
                        return;
                    }

                    // 3. READ CHECKBOX: Do nothing, let checkbox renderer handle it
                    if (colId === 'Read') {
                        return;
                    }

                    // 4. ROW CLICK (all other columns): Navigate to detail view
                    // Select this row (clearing others) to trigger navigation
                    event.node.setSelected(true, true);
                }
            """)

            # Custom CSS for professional appearance (theme-aware)
            # Structural properties are identical in both themes, only colors differ
            if st.session_state.theme == 'dark':
                custom_css = {
                    # Headers - consistent structure, dark colors
                    ".ag-header-cell-label": {
                        "font-weight": "600",
                        "font-size": "14px",
                        "font-family": "inherit",
                        "color": "#FFFFFF !important"
                    },
                    ".ag-header-cell-text": {
                        "color": "#FFFFFF !important",
                        "font-size": "14px"
                    },
                    ".ag-header": {
                        "background-color": "#262730 !important",
                        "border-bottom": "1px solid #444444 !important",
                        "height": "40px !important"
                    },
                    ".ag-header-cell": {
                        "background-color": "#262730 !important",
                        "color": "#FFFFFF !important",
                        "padding": "0 8px !important"
                    },
                    # Rows and cells - consistent structure, dark colors
                    ".ag-root-wrapper": {
                        "background-color": "#1E1E1E !important",
                        "width": "100% !important",
                        "font-family": "inherit",
                        "font-size": "14px"
                    },
                    ".ag-row": {
                        "border-bottom": "1px solid #444444 !important",
                        "background-color": "#1E1E1E !important",
                        "height": "60px !important"
                    },
                    ".ag-cell": {
                        "padding": "8px !important",
                        "display": "flex !important",
                        "align-items": "center !important",
                        "color": "#E0E0E0 !important",
                        "overflow": "hidden !important",
                        "text-overflow": "ellipsis !important",
                        "font-size": "14px !important",
                        "font-family": "inherit !important",
                        "line-height": "1.5 !important"
                    },
                    # Multi-line text cells with line-clamp
                    ".ag-cell .ag-cell-value": {
                        "display": "-webkit-box !important",
                        "-webkit-line-clamp": "2 !important",
                        "-webkit-box-orient": "vertical !important",
                        "overflow": "hidden !important",
                        "text-overflow": "ellipsis !important",
                        "white-space": "normal !important",
                        "line-height": "1.4 !important",
                        "max-height": "42px !important",
                        "width": "100% !important"
                    },
                    ".ag-row-hover": {"background-color": "#2D2D2D !important"},
                    ".ag-row-hover .doi-edit-icon": {"opacity": "1 !important"},
                    # Grid background - full width
                    ".ag-center-cols-viewport": {"background-color": "#1E1E1E !important"},
                    ".ag-body-viewport": {"background-color": "#1E1E1E !important"},
                }
            else:
                custom_css = {
                    # Headers - consistent structure, light colors
                    ".ag-header-cell-label": {
                        "font-weight": "600",
                        "font-size": "14px",
                        "font-family": "inherit",
                        "color": "#2c3e50 !important"
                    },
                    ".ag-header-cell-text": {
                        "color": "#2c3e50 !important",
                        "font-size": "14px"
                    },
                    ".ag-header": {
                        "background-color": "#F0F2F6 !important",
                        "border-bottom": "1px solid #D0D0D0 !important",
                        "height": "40px !important"
                    },
                    ".ag-header-cell": {
                        "background-color": "#F0F2F6 !important",
                        "color": "#2c3e50 !important",
                        "padding": "0 8px !important"
                    },
                    # Rows and cells - consistent structure, light colors
                    ".ag-root-wrapper": {
                        "background-color": "#FFFFFF !important",
                        "width": "100% !important",
                        "font-family": "inherit",
                        "font-size": "14px"
                    },
                    ".ag-row": {
                        "border-bottom": "1px solid #ecf0f1 !important",
                        "background-color": "#FFFFFF !important",
                        "height": "60px !important"
                    },
                    ".ag-cell": {
                        "padding": "8px !important",
                        "display": "flex !important",
                        "align-items": "center !important",
                        "color": "#262730 !important",
                        "overflow": "hidden !important",
                        "text-overflow": "ellipsis !important",
                        "font-size": "14px !important",
                        "font-family": "inherit !important",
                        "line-height": "1.5 !important"
                    },
                    # Multi-line text cells with line-clamp
                    ".ag-cell .ag-cell-value": {
                        "display": "-webkit-box !important",
                        "-webkit-line-clamp": "2 !important",
                        "-webkit-box-orient": "vertical !important",
                        "overflow": "hidden !important",
                        "text-overflow": "ellipsis !important",
                        "white-space": "normal !important",
                        "line-height": "1.4 !important",
                        "max-height": "42px !important",
                        "width": "100% !important"
                    },
                    ".ag-row-hover": {"background-color": "#f8f9fa !important"},
                    ".ag-row-hover .doi-edit-icon": {"opacity": "1 !important"},
                    # Grid background - full width
                    ".ag-center-cols-viewport": {"background-color": "#FFFFFF !important"},
                    ".ag-body-viewport": {"background-color": "#FFFFFF !important"},
                }

            # Use streamlit theme for both modes - consistent base, colors controlled by custom CSS
            ag_theme = 'streamlit'

            # Use key to preserve grid state across reruns
            grid_key = f"library_grid_{st.session_state.theme}"

            grid_response = AgGrid(
                df,
                gridOptions=grid_options,
                update_mode=GridUpdateMode.SELECTION_CHANGED,  # Update on selection changes for row click navigation
                fit_columns_on_grid_load=True,  # Auto-fit columns to fill container width
                theme=ag_theme,
                custom_css=custom_css,
                allow_unsafe_jscode=True,
                enable_enterprise_modules=False,
                height=1400,  # Fixed height to show ~23 rows with internal scrolling
                reload_data=False,  # Improve performance by not reloading data unnecessarily
                key=grid_key  # Preserve state across reruns
            )

            # Handle delete button click
            if delete_button and grid_response['selected_rows'] is not None and len(grid_response['selected_rows']) > 0:
                selected_rows_df = pd.DataFrame(grid_response['selected_rows'])

                # Check if user has disabled confirmation dialogs
                settings = load_settings()
                skip_delete_confirmation = settings.get('skip_delete_confirmation', False)

                if skip_delete_confirmation:
                    # Delete immediately without confirmation
                    success_count = 0
                    for _, row in selected_rows_df.iterrows():
                        result = soft_delete_paper(row['_filename'])
                        if result['success']:
                            success_count += 1

                    st.session_state.reload_papers = True  # Invalidate cache
                    st.toast(f"🗑️ Moved {success_count} paper(s) to trash", icon="✅")
                    st.rerun()
                else:
                    # Store papers to delete in session state and show dialog
                    st.session_state.papers_to_delete = selected_rows_df
                    st.session_state.show_delete_dialog = True
                    st.rerun()

            # Handle confirmed deletion (outside dialog to ensure clean close)
            if st.session_state.get('delete_confirmed', False):
                papers_to_process = st.session_state.get('papers_to_delete_confirmed', None)
                if papers_to_process is not None:
                    success_count = 0
                    for _, row in papers_to_process.iterrows():
                        result = soft_delete_paper(row['_filename'])
                        if result['success']:
                            success_count += 1

                    st.session_state.reload_papers = True  # Invalidate cache
                    st.toast(f"🗑️ Moved {success_count} paper(s) to trash", icon="✅")

                    # Clear the confirmation flags
                    st.session_state.delete_confirmed = False
                    if 'papers_to_delete_confirmed' in st.session_state:
                        del st.session_state.papers_to_delete_confirmed
                    st.rerun()

            # Show delete confirmation dialog
            if st.session_state.get('show_delete_dialog', False) and 'papers_to_delete' in st.session_state:
                @st.dialog("⚠️ Confirm Delete")
                def confirm_delete_dialog():
                    papers_df = st.session_state.papers_to_delete
                    paper_titles = [row['_paper_title'] for _, row in papers_df.iterrows()]
                    num_papers = len(paper_titles)

                    st.write(f"Are you sure you want to delete **{num_papers} paper(s)**?")
                    st.write("")
                    st.write("**Papers to be deleted:**")
                    for title in paper_titles[:5]:  # Show first 5
                        st.write(f"• {title}")
                    if num_papers > 5:
                        st.write(f"_... and {num_papers - 5} more_")

                    st.write("")

                    # Don't ask again checkbox
                    dont_ask = st.checkbox("Don't ask again (skip confirmation in the future)", key="dont_ask_delete")

                    st.write("")
                    col1, col2 = st.columns(2)

                    with col1:
                        if st.button("✓ Confirm Delete", type="primary", use_container_width=True):
                            # Save preference if checkbox is checked
                            if dont_ask:
                                settings = load_settings()
                                settings['skip_delete_confirmation'] = True
                                save_settings(settings)

                            # Store papers for deletion and set confirmation flag
                            st.session_state.papers_to_delete_confirmed = papers_df.copy()
                            st.session_state.delete_confirmed = True

                            # Close dialog
                            st.session_state.show_delete_dialog = False
                            if 'papers_to_delete' in st.session_state:
                                del st.session_state.papers_to_delete
                            st.rerun()

                    with col2:
                        if st.button("✗ Cancel", use_container_width=True):
                            # Clear session state and close dialog
                            st.session_state.show_delete_dialog = False
                            if 'papers_to_delete' in st.session_state:
                                del st.session_state.papers_to_delete
                            st.rerun()

                confirm_delete_dialog()

            # Handle read status changes
            if grid_response['data'] is not None:
                updated_df = pd.DataFrame(grid_response['data'])
                for idx, row in updated_df.iterrows():
                    original_status = df.iloc[idx]['Read']
                    new_status = row['Read']
                    if original_status != new_status:
                        filename = row['_filename']
                        if new_status:
                            read_status.mark_as_read(filename)
                        else:
                            read_status.mark_as_unread(filename)
                        st.rerun()
                        break

            # Handle row click navigation
            # Only navigate if exactly one row is selected (indicates row click, not checkbox multi-select)
            # and delete button wasn't clicked
            if not delete_button and grid_response.get('selected_rows') is not None:
                selected_rows = grid_response['selected_rows']
                if len(selected_rows) == 1:
                    # Single row selected - navigate to detail page
                    selected_row = pd.DataFrame(selected_rows).iloc[0]
                    filename = selected_row['_filename']
                    st.session_state.selected_paper = filename
                    st.rerun()

    # ===== TAB 3: DISCOVER =====
    with tab3:
        st.session_state.active_tab = "Discover"

        st.markdown("### 🔍 Discover")
        st.caption("Search for papers and discover what you're missing")

        st.markdown("---")

        # Section 1: Search the Field
        st.markdown("#### 🔎 Search the Field")
        st.caption("Search for papers across all of academia using Semantic Scholar")

        # Check API key status and show warning if needed
        ss_api_key = semantic_scholar.get_api_key()
        if not ss_api_key:
            st.warning("⚠️ **Using unauthenticated access (100 searches per 5 min).** Add a [free API key](https://www.semanticscholar.org/product/api) in Settings for 5,000 searches per 5 min.")
        else:
            st.info("✓ API key active - Higher rate limits enabled (5,000 searches per 5 min)")

        # Search input
        search_col1, search_col2 = st.columns([4, 1])
        with search_col1:
            search_query = st.text_input(
                "Search for papers",
                placeholder="e.g., EIS SOH estimation LFP, machine learning battery degradation",
                key="semantic_scholar_search",
                label_visibility="collapsed"
            )
        with search_col2:
            sort_by = st.selectbox(
                "Sort by",
                options=["Relevance", "Citations"],
                key="semantic_scholar_sort",
                label_visibility="collapsed"
            )

        # Search button
        if st.button("🔍 Search", type="primary", use_container_width=False):
            if search_query.strip():
                # Create cache key from query + sort
                new_query = search_query.strip()
                new_sort = "citationCount" if sort_by == "Citations" else "relevance"
                cache_key = f"{new_query}|{new_sort}"

                # Only search if query/sort changed
                if st.session_state.get('ss_cache_key') != cache_key:
                    st.session_state['ss_search_query'] = new_query
                    st.session_state['ss_search_sort'] = new_sort
                    st.session_state['ss_search_results'] = None  # Clear previous results
                    st.session_state['ss_cache_key'] = cache_key
                    st.session_state['ss_search_triggered'] = True
            else:
                st.warning("Please enter a search query")

        # Display search results
        if st.session_state.get('ss_search_query') and st.session_state.get('ss_search_triggered'):
            query = st.session_state['ss_search_query']
            sort_type = st.session_state.get('ss_search_sort', 'relevance')

            # Perform search if not already cached
            if st.session_state.get('ss_search_results') is None:
                with st.spinner(f"Searching Semantic Scholar for '{query}'..."):
                    # Add a small delay before the API call to help with rate limiting
                    time.sleep(1.5)

                    result = semantic_scholar.search_papers(
                        query=query,
                        limit=20,
                        sort=sort_type
                    )
                    st.session_state['ss_search_results'] = result

            result = st.session_state['ss_search_results']

            if result['success']:
                papers_raw = result['data']
                total = result['total']

                if papers_raw:
                    # Format papers
                    papers_formatted = [semantic_scholar.format_paper_for_display(p) for p in papers_raw]

                    # Build library lookup sets
                    from lib.gap_analysis import normalize_doi, normalize_title
                    library_dois = set()
                    library_titles = set()

                    for paper in papers:
                        if paper.get('doi'):
                            library_dois.add(normalize_doi(paper['doi']))
                        if paper.get('title'):
                            library_titles.add(normalize_title(paper['title']))

                    # Check which papers are in library
                    papers_formatted = semantic_scholar.check_papers_in_library(
                        papers_formatted,
                        library_dois,
                        library_titles
                    )

                    st.caption(f"Found {total:,} papers. Showing top {len(papers_formatted)} results.")

                    # Prepare AG Grid data
                    grid_data = []
                    for i, paper in enumerate(papers_formatted, 1):
                        grid_data.append({
                            'Rank': i,
                            'Title': paper['title'],
                            'Authors': paper['authors'],
                            'Year': paper['year'],
                            'Citations': paper['citation_count'],
                            'In Library': '✓' if paper['in_library'] else '—',
                            '_doi': paper['doi'],
                            '_pdf_url': paper['pdf_url'],
                            '_is_open_access': paper['is_open_access'],
                            '_in_library': paper['in_library'],
                            '_full_title': paper['title'],
                            '_full_authors': paper['authors'],
                            '_abstract': paper['abstract'],
                            '_authors_list': paper['authors_list'],
                            '_journal': paper['journal'],
                            '_venue': paper['venue']
                        })

                    df_search = pd.DataFrame(grid_data)

                    # Configure AG Grid
                    gb_search = GridOptionsBuilder.from_dataframe(df_search)

                    # Rank column
                    gb_search.configure_column("Rank",
                        width=70,
                        cellStyle={'textAlign': 'center', 'fontWeight': 'bold'}
                    )

                    # Title column
                    gb_search.configure_column("Title",
                        flex=3,
                        minWidth=300,
                        wrapText=True,
                        autoHeight=False,
                        cellStyle={
                            'whiteSpace': 'normal !important',
                            'lineHeight': '1.4 !important',
                            'display': '-webkit-box !important',
                            '-webkit-line-clamp': '2 !important',
                            '-webkit-box-orient': 'vertical !important',
                            'overflow': 'hidden !important',
                            'textOverflow': 'ellipsis !important',
                            'padding': '8px !important',
                            'maxHeight': '50px !important'
                        },
                        tooltipField="_full_title"
                    )

                    # Authors column
                    gb_search.configure_column("Authors",
                        flex=2,
                        minWidth=150,
                        wrapText=True,
                        cellStyle={
                            'whiteSpace': 'normal !important',
                            'overflow': 'hidden !important',
                            'textOverflow': 'ellipsis !important',
                            'padding': '8px !important'
                        },
                        tooltipField="_full_authors"
                    )

                    # Year column
                    gb_search.configure_column("Year",
                        width=70,
                        cellStyle={'textAlign': 'center'}
                    )

                    # Citations column
                    gb_search.configure_column("Citations",
                        width=90,
                        cellStyle={'textAlign': 'center', 'fontWeight': 'bold', 'color': '#0066cc'}
                    )

                    # In Library column
                    gb_search.configure_column("In Library",
                        width=100,
                        cellStyle={'textAlign': 'center', 'fontSize': '16px'}
                    )

                    # Hide internal columns
                    for col in ['_doi', '_pdf_url', '_is_open_access', '_in_library', '_full_title',
                               '_full_authors', '_abstract', '_authors_list', '_journal', '_venue']:
                        gb_search.configure_column(col, hide=True)

                    # Grid options
                    gb_search.configure_selection(selection_mode="single", use_checkbox=False)
                    gb_search.configure_grid_options(
                        domLayout='normal',
                        rowHeight=50,
                        enableCellTextSelection=True,
                        ensureDomOrder=True,
                        suppressRowClickSelection=False
                    )

                    grid_options_search = gb_search.build()

                    # Display AG Grid
                    grid_response_search = AgGrid(
                        df_search,
                        gridOptions=grid_options_search,
                        height=500,
                        theme="streamlit",
                        update_mode=GridUpdateMode.SELECTION_CHANGED,
                        allow_unsafe_jscode=True,
                        fit_columns_on_grid_load=True
                    )

                    # Show selected paper details
                    selected_rows_search = grid_response_search.get('selected_rows', None)
                    if selected_rows_search is not None and len(selected_rows_search) > 0:
                        selected_search = selected_rows_search[0]

                        st.markdown("---")
                        st.markdown(f"### Selected: {selected_search['Title']}")

                        col1_search, col2_search = st.columns([3, 1])

                        with col1_search:
                            st.markdown(f"**Authors:** {selected_search['_full_authors']}")
                            st.markdown(f"**Year:** {selected_search['Year']}")

                            if selected_search['_journal']:
                                st.markdown(f"**Journal:** {selected_search['_journal']}")
                            elif selected_search['_venue']:
                                st.markdown(f"**Venue:** {selected_search['_venue']}")

                            st.markdown(f"**Citations:** {selected_search['Citations']}")

                            # Abstract
                            if selected_search['_abstract']:
                                with st.expander("📄 Abstract", expanded=False):
                                    st.caption(selected_search['_abstract'])

                        with col2_search:
                            # DOI
                            doi_search = selected_search['_doi']
                            if doi_search:
                                st.markdown(f"**DOI:** [{doi_search}](https://doi.org/{doi_search})")

                            # In library status
                            if selected_search['_in_library']:
                                st.success("**✓ In Library**")
                            else:
                                # Add to Library button
                                pdf_url = selected_search['_pdf_url']
                                is_open_access = selected_search['_is_open_access']

                                button_label = "📥 Add to Library"
                                if is_open_access and pdf_url:
                                    button_label = "📥 Add + Download PDF"

                                if st.button(button_label, type="primary", use_container_width=True, key="import_ss_selected"):
                                    # Import paper
                                    with st.spinner("Importing paper..."):
                                        # Create safe filename
                                        safe_title = re.sub(r'[^\w\s-]', '', selected_search['Title'][:50])
                                        safe_title = re.sub(r'[-\s]+', '_', safe_title)

                                        # Prepare metadata
                                        metadata = {
                                            'title': selected_search['Title'],
                                            'authors': selected_search['_authors_list'],
                                            'year': selected_search['Year'],
                                            'journal': selected_search['_journal'] or selected_search['_venue'],
                                            'doi': doi_search,
                                            'abstract': selected_search['_abstract'],
                                            'author_keywords': []
                                        }

                                        # Download PDF if available
                                        pdf_downloaded = False
                                        filename = None

                                        if is_open_access and pdf_url:
                                            st.info("📥 Downloading open access PDF...")
                                            filename = f"{safe_title}.pdf"
                                            pdf_path = Path("papers") / filename

                                            download_result = semantic_scholar.download_pdf(pdf_url, pdf_path)

                                            if download_result['success']:
                                                pdf_downloaded = True
                                                st.success("✓ PDF downloaded!")
                                            else:
                                                st.warning(f"Could not download PDF: {download_result['message']}")
                                                st.info("Adding as metadata-only...")

                                        # Save metadata
                                        if not filename:
                                            # Metadata-only
                                            if doi_search:
                                                safe_doi = doi_search.replace('/', '_').replace('.', '_')
                                                filename = f"doi_{safe_doi}.pdf"
                                            else:
                                                filename = f"{safe_title}.pdf"

                                        try:
                                            filename_saved = save_metadata_only_paper(doi_search if doi_search else "", metadata)

                                            if pdf_downloaded:
                                                st.success("✓ Paper added with PDF!")
                                            else:
                                                st.success("✓ Paper added as metadata-only!")

                                            st.info("View it in the Library tab.")
                                            time.sleep(2)
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"Failed to save paper: {e}")

                else:
                    st.info(f"No results found for '{query}'")

            else:
                st.error(f"Search failed: {result['error']}")

        # Section 2: Frequently Cited Papers You Don't Have
        st.markdown("---")
        st.markdown("#### 📚 Frequently Cited Papers You Don't Have")
        st.caption("Papers frequently cited in your library that you haven't added yet")

        # Get gap analysis data
        with st.spinner("Analyzing references across your library..."):
            top_gaps = gap_analysis.get_top_gaps(n=20)

        if not top_gaps:
            st.info("No frequently cited papers found! Either you have all referenced papers, or no references have been extracted yet.")
        else:
            # Prepare data for AG Grid
            grid_data = []
            for i, gap in enumerate(top_gaps, 1):
                # Format authors (truncate if too long)
                authors_display = gap['authors'][:60] + '...' if len(gap['authors']) > 60 else gap['authors']

                grid_data.append({
                    'Rank': i,
                    'Title': gap['title'],
                    'Authors': authors_display,
                    'Year': gap['year'] if gap['year'] else '—',
                    'Cited By': gap['citation_count'],
                    '_doi': gap['doi'],  # Hidden column for import
                    '_full_title': gap['title'],  # For tooltip
                    '_full_authors': gap['authors'],  # For tooltip
                    '_cited_by_list': gap['cited_by']  # For modal/details
                })

            df = pd.DataFrame(grid_data)

            # Configure AG Grid
            gb = GridOptionsBuilder.from_dataframe(df)

            # Rank column
            gb.configure_column("Rank",
                width=80,
                minWidth=60,
                maxWidth=100,
                cellStyle={'textAlign': 'center', 'fontWeight': 'bold'}
            )

            # Title column
            gb.configure_column("Title",
                flex=3,
                minWidth=300,
                wrapText=True,
                autoHeight=False,
                cellStyle={
                    'whiteSpace': 'normal !important',
                    'lineHeight': '1.4 !important',
                    'display': '-webkit-box !important',
                    '-webkit-line-clamp': '2 !important',
                    '-webkit-box-orient': 'vertical !important',
                    'overflow': 'hidden !important',
                    'textOverflow': 'ellipsis !important',
                    'padding': '8px !important',
                    'maxHeight': '50px !important'
                },
                tooltipField="_full_title"
            )

            # Authors column
            gb.configure_column("Authors",
                flex=2,
                minWidth=180,
                wrapText=True,
                cellStyle={
                    'whiteSpace': 'normal !important',
                    'lineHeight': '1.4 !important',
                    'overflow': 'hidden !important',
                    'textOverflow': 'ellipsis !important',
                    'padding': '8px !important'
                },
                tooltipField="_full_authors"
            )

            # Year column
            gb.configure_column("Year",
                width=80,
                minWidth=70,
                maxWidth=100,
                cellStyle={'textAlign': 'center'}
            )

            # Cited By column
            gb.configure_column("Cited By",
                headerName="Cited By",
                width=100,
                minWidth=90,
                maxWidth=120,
                cellStyle={
                    'textAlign': 'center',
                    'fontWeight': 'bold',
                    'color': '#0066cc'
                }
            )

            # Hide internal columns
            gb.configure_column("_doi", hide=True)
            gb.configure_column("_full_title", hide=True)
            gb.configure_column("_full_authors", hide=True)
            gb.configure_column("_cited_by_list", hide=True)

            # Grid options - enable multi-select with checkboxes
            gb.configure_selection(selection_mode="multiple", use_checkbox=True)
            gb.configure_grid_options(
                domLayout='normal',
                rowHeight=50,
                enableCellTextSelection=True,
                ensureDomOrder=True,
                suppressRowClickSelection=True  # Don't select on row click, only checkbox
            )

            grid_options = gb.build()

            # Action buttons
            st.caption("💡 **Tip:** Check boxes to select papers, then use buttons below")
            btn_col1, btn_col2, spacer_col = st.columns([1, 1, 3])
            with btn_col1:
                add_selected_button = st.button("📥 Add Selected", type="primary", use_container_width=True, key="add_selected_gaps")
            with btn_col2:
                add_all_button = st.button("📥 Add All", type="secondary", use_container_width=True, key="add_all_gaps")

            # Display AG Grid
            grid_response = AgGrid(
                df,
                gridOptions=grid_options,
                height=600,
                theme="streamlit",
                update_mode=GridUpdateMode.SELECTION_CHANGED,
                allow_unsafe_jscode=True,
                fit_columns_on_grid_load=True
            )

            # Get selected rows
            selected_rows = grid_response.get('selected_rows', None)

            # Handle Add Selected button
            if add_selected_button:
                if selected_rows is not None and len(selected_rows) > 0:
                    with st.spinner(f"Adding {len(selected_rows)} paper(s)..."):
                        results = []
                        for row in selected_rows:
                            doi = row.get('_doi', '')
                            title = row.get('Title', 'Unknown')

                            # Try to add paper with PDF search
                            result = add_paper_with_pdf_search(doi, title, row.get('_full_authors', ''), row.get('Year', ''), url='')
                            results.append(result)

                        # Show summary
                        success_count = sum(1 for r in results if r['success'])
                        st.success(f"✓ Added {success_count} of {len(selected_rows)} paper(s)")

                        if success_count > 0:
                            time.sleep(2)
                            st.rerun()
                else:
                    st.warning("Please select papers to add (use checkboxes)")

            # Handle Add All button
            if add_all_button:
                if st.session_state.get('confirm_add_all_gaps', False):
                    with st.spinner(f"Adding all {len(top_gaps)} papers..."):
                        results = []
                        for gap in top_gaps:
                            doi = gap.get('doi', '')
                            title = gap.get('title', 'Unknown')

                            result = add_paper_with_pdf_search(doi, title, gap.get('authors', ''), gap.get('year', ''), url='')
                            results.append(result)

                        success_count = sum(1 for r in results if r['success'])
                        st.success(f"✓ Added {success_count} of {len(top_gaps)} paper(s)")
                        st.session_state['confirm_add_all_gaps'] = False

                        if success_count > 0:
                            time.sleep(2)
                            st.rerun()
                else:
                    st.warning(f"⚠️ Click again to confirm adding all {len(top_gaps)} papers")
                    st.session_state['confirm_add_all_gaps'] = True

    # ===== TAB 4: RESEARCH =====
    with tab4:
        st.session_state.active_tab = "Research"

        # Prominent "Ask a Question" section at the top
        st.subheader("Ask a Research Question")
        st.caption("Ask questions about your battery research papers using AI-powered search")

        question = st.text_area(
            "Your question:",
            placeholder="What factors affect battery degradation?",
            height=100,
            label_visibility="collapsed",
            key="research_question"
        )

        # Filters in horizontal layout
        st.caption("**Optional Filters:**")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            filter_chemistry = st.selectbox(
                "Chemistry",
                options=["All"] + filter_options['chemistries'],
                key="research_filter_chemistry"
            )
            filter_chemistry = None if filter_chemistry == "All" else filter_chemistry
        with col2:
            filter_topic = st.selectbox(
                "Topic",
                options=["All"] + filter_options['topics'],
                key="research_filter_topic"
            )
            filter_topic = None if filter_topic == "All" else filter_topic
        with col3:
            filter_paper_type = st.selectbox(
                "Paper Type",
                options=["All"] + filter_options['paper_types'],
                key="research_filter_paper_type"
            )
            filter_paper_type = None if filter_paper_type == "All" else filter_paper_type
        with col4:
            # Collection filter
            all_collections = collections.get_all_collections()
            collection_options = ["All Papers"] + [c['name'] for c in all_collections]
            filter_collection = st.selectbox(
                "📁 Collection",
                options=collection_options,
                key="research_filter_collection"
            )
            # Get collection filenames if a specific collection is selected
            filter_collection_filenames = None
            if filter_collection != "All Papers":
                selected_collection = next((c for c in all_collections if c['name'] == filter_collection), None)
                if selected_collection:
                    filter_collection_filenames = set(collections.get_collection_papers(selected_collection['id']))

        # Search button
        if st.button("Search", type="primary", use_container_width=True, key="research_search_button"):
            if not question:
                st.warning("Please enter a question")
            else:
                api_key = get_api_key()
                if api_key:
                    # Show progress with improved retrieval steps
                    progress_text = st.empty()
                    progress_bar = st.progress(0)

                    try:
                        # Step 1: Query expansion
                        progress_text.text("Step 1/4: Expanding query...")
                        progress_bar.progress(25)

                        # Step 2: Hybrid search
                        progress_text.text("Step 2/4: Hybrid search (vector + BM25)...")
                        progress_bar.progress(50)

                        # Step 3: Reranking
                        progress_text.text("Step 3/4: Reranking by relevance...")
                        progress_bar.progress(75)

                        # Use improved retrieval pipeline
                        chunks = rag.retrieve_with_hybrid_and_reranking(
                            question=question,
                            api_key=api_key,
                            top_k=5,
                            n_candidates=15,
                            alpha=0.5,
                            filter_chemistry=filter_chemistry,
                            filter_topic=filter_topic,
                            filter_paper_type=filter_paper_type,
                            filter_collection_filenames=filter_collection_filenames,
                            enable_query_expansion=True,
                            enable_reranking=True
                        )

                        if not chunks:
                            progress_text.empty()
                            progress_bar.empty()
                            st.warning("No relevant passages found. Try removing filters.")
                        else:
                            # Step 4: Query Claude
                            progress_text.text("Step 4/4: Generating answer with Claude...")
                            progress_bar.progress(90)

                            # Use backend for LLM query
                            answer = rag.query_claude(question, chunks, api_key)

                            progress_bar.progress(100)
                            progress_text.empty()
                            progress_bar.empty()

                            # Prepare query result
                            query_result = {
                                'question': question,
                                'answer': answer,
                                'chunks': chunks,
                                'filters': {
                                    'chemistry': filter_chemistry,
                                    'topic': filter_topic,
                                    'paper_type': filter_paper_type
                                }
                            }

                            # Save to history database
                            query_history.save_query(
                                question=question,
                                answer=answer,
                                chunks=chunks,
                                filters=query_result['filters']
                            )

                            # Save to session state
                            st.session_state.query_result = query_result
                            st.rerun()
                    except RuntimeError as e:
                        progress_text.empty()
                        progress_bar.empty()
                        st.error(f"Error: {e}")

        st.divider()

        # Show results if available
        if st.session_state.query_result:
            result = st.session_state.query_result

            # Show indicator if loaded from history
            if result.get('from_history'):
                st.info("Viewing query from history")

            # Show question
            st.subheader("Question")
            st.info(result['question'])

            # Show active filters
            if any(result['filters'].values()):
                st.caption("**Active Filters:** " +
                          ", ".join([f"{k}: {v}" for k, v in result['filters'].items() if v]))

            st.divider()

            # Show answer with clickable document references
            st.subheader("Answer")

            # Convert "Document X" references to clickable anchor links
            answer_html = result['answer']
            # Match patterns like "Document 1", "Document 2, page 5", etc.
            pattern = r'(Document\s+(\d+)(?:,\s*page\s+\d+)?)'
            def make_link(match):
                full_text = match.group(1)
                doc_num = match.group(2)
                return f'<a href="#passage-{doc_num}" style="color: #1f77b4; text-decoration: none;">{full_text}</a>'

            answer_html = re.sub(pattern, make_link, answer_html)
            st.markdown(answer_html, unsafe_allow_html=True)

            st.divider()

            # Show sources with clickable links to detail pages
            st.subheader("Sources & Citations")

            # Get unique papers cited
            cited_papers = {}
            for chunk in result['chunks']:
                filename = chunk['filename']
                if filename not in cited_papers:
                    cited_papers[filename] = []
                cited_papers[filename].append(chunk['page_num'])

            st.write(f"**{len(cited_papers)} papers cited:**")

            # Make paper names clickable buttons to navigate to detail page
            for filename, pages in cited_papers.items():
                pages_str = ', '.join(map(str, sorted(set(pages))))
                col1, col2 = st.columns([3, 1])
                with col1:
                    if st.button(f"📄 {filename}", key=f"source_link_{filename}", use_container_width=True):
                        st.session_state.selected_paper = filename
                        st.rerun()
                with col2:
                    st.caption(f"Pages: {pages_str}")

            st.divider()

            # Show chunks with anchor IDs
            st.subheader("Retrieved Passages")
            for i, chunk in enumerate(result['chunks'], 1):
                section_label = f" - {chunk['section_name']}" if chunk.get('section_name') and chunk['section_name'] != 'Content' else ""

                # Add anchor ID to the passage
                st.markdown(f'<div id="passage-{i}"></div>', unsafe_allow_html=True)

                with st.expander(f"Passage {i}: {chunk['filename']} (page {chunk['page_num']}){section_label}"):
                    st.write(chunk['text'])
                    metadata_parts = []
                    if chunk.get('section_name'):
                        metadata_parts.append(f"Section: {chunk['section_name']}")
                    if chunk['chemistries']:
                        metadata_parts.append(f"Chemistries: {', '.join(chunk['chemistries'])}")
                    if chunk['topics']:
                        metadata_parts.append(f"Topics: {', '.join(chunk['topics'][:5])}")
                    if metadata_parts:
                        st.caption(" | ".join(metadata_parts))

        else:
            st.info("💡 **Tip:** Ask questions about your battery research papers and get AI-powered answers")
            st.write("**Example questions:**")
            st.write("- What factors affect battery degradation?")
            st.write("- How does temperature impact NMC vs LFP cells?")
            st.write("- What is lithium plating and when does it occur?")
            st.write("- How to estimate state of health?")

    # ===== TAB 5: HISTORY =====
    with tab5:
        st.session_state.active_tab = "History"

        st.subheader("Query History")

        # Get all queries
        all_queries = query_history.get_all_queries()

        if not all_queries:
            st.info("No query history yet. Run a query to see it saved here!")
        else:
            # Show starred queries first if any exist
            starred_queries = [q for q in all_queries if q['is_starred']]
            unstarred_queries = [q for q in all_queries if not q['is_starred']]

            # Filter options
            col1, col2, col3 = st.columns([2, 1, 1])
            with col1:
                show_filter = st.selectbox(
                    "Show:",
                    ["All Queries", "Starred Only", "Recent (Last 10)"],
                    key="history_filter"
                )
            with col2:
                st.metric("Total Queries", len(all_queries))
            with col3:
                st.metric("Starred", len(starred_queries))

            # Apply filter
            if show_filter == "Starred Only":
                queries_to_show = starred_queries
            elif show_filter == "Recent (Last 10)":
                queries_to_show = all_queries[:10]
            else:
                queries_to_show = all_queries

            st.divider()

            # Show starred queries section
            if starred_queries and show_filter == "All Queries":
                st.subheader("⭐ Starred Queries")
                for query in starred_queries:
                    display_query_card(query)
                    st.divider()

                if unstarred_queries:
                    st.subheader("📋 All Queries")

            # Show queries
            if not queries_to_show:
                st.info("No queries match the selected filter.")
            else:
                for query in queries_to_show:
                    if show_filter != "All Queries" or not query['is_starred']:
                        display_query_card(query)
                        st.divider()

            # Clear all history button (at the bottom, with warning)
            st.divider()
            with st.expander("⚠️ Danger Zone", expanded=False):
                st.warning("Clear all query history? This cannot be undone!")
                if st.button("🗑️ Clear All History", type="secondary"):
                    count = query_history.clear_all_history()
                    st.success(f"Deleted {count} queries from history")
                    st.rerun()

    # ===== TAB 6: SETTINGS =====
    with tab6:
        st.session_state.active_tab = "Settings"

        st.markdown("### ⚙️ Application Settings")

        # Theme Settings
        st.markdown("#### Appearance")

        current_theme = st.session_state.theme
        theme_label = "🌙 Switch to Dark Mode" if current_theme == "light" else "☀️ Switch to Light Mode"

        if st.button(theme_label, use_container_width=True, type="primary"):
            # Toggle theme
            new_theme = "dark" if current_theme == "light" else "light"
            st.session_state.theme = new_theme
            save_theme_preference(new_theme)
            st.rerun()

        st.caption(f"Current theme: **{current_theme.title()}**")

        st.divider()

        # Backup & Restore
        st.markdown("#### Backup & Restore")
        st.caption("Create backups of your database and restore from previous backups.")

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**Create Backup**")
            if st.button("📦 Create Backup", use_container_width=True):
                with st.spinner("Creating backup..."):
                    from lib import backup as backup_module
                    result = backup_module.create_backup(include_logs=False)

                    if result['success']:
                        st.success(f"✓ Backup created!")
                        st.caption(f"Size: {result['size_mb']} MB")
                        st.caption(f"Files: {result['file_count']}")

                        # Offer download
                        with open(result['backup_path'], 'rb') as f:
                            st.download_button(
                                label="💾 Download Backup",
                                data=f,
                                file_name=Path(result['backup_path']).name,
                                mime="application/zip",
                                use_container_width=True
                            )
                    else:
                        st.error(f"Backup failed: {result.get('error', 'Unknown error')}")

        with col2:
            st.markdown("**Restore from Backup**")
            # Show existing backups
            from lib import backup as backup_module
            backups = backup_module.list_backups()

            if backups:
                backup_options = {f"{b['name']} ({b['size_mb']} MB)": b['path']
                                  for b in backups}

                selected = st.selectbox(
                    "Select backup to restore:",
                    options=list(backup_options.keys()),
                    key="restore_backup_select",
                    label_visibility="collapsed"
                )

                if st.button("♻️ Restore", type="secondary", use_container_width=True):
                    backup_path = Path(backup_options[selected])

                    # Confirmation
                    if st.session_state.get('restore_confirmed', False):
                        with st.spinner("Restoring backup..."):
                            result = backup_module.restore_backup(backup_path)

                            if result['success']:
                                st.success(result['message'])
                                st.info("Please refresh the page to see restored data.")
                                st.session_state['restore_confirmed'] = False
                            else:
                                st.error(f"Restore failed: {result.get('error')}")
                    else:
                        st.warning("⚠️ This will replace current data. Click again to confirm.")
                        st.session_state['restore_confirmed'] = True
            else:
                st.info("No backups available yet")

        st.divider()

        # Delete Confirmation Settings
        st.markdown("#### Delete Confirmation")
        st.caption("Control whether to show confirmation dialog when deleting papers.")

        settings = load_settings()
        skip_confirmation = settings.get('skip_delete_confirmation', False)

        if skip_confirmation:
            st.info("✓ Delete confirmation is currently **disabled**. Papers will be deleted immediately.")
            if st.button("Enable Delete Confirmation", use_container_width=True):
                settings['skip_delete_confirmation'] = False
                save_settings(settings)
                st.success("Delete confirmation enabled!")
                st.rerun()
        else:
            st.info("✓ Delete confirmation is currently **enabled**. You'll see a popup before deleting.")
            st.caption("You can disable it by checking 'Don't ask again' in the delete confirmation dialog.")

        st.divider()

        # Semantic Scholar API Key Settings
        st.markdown("#### 🔑 Semantic Scholar API Key")
        st.caption("Add your API key for higher rate limits when searching papers (5,000 vs 100 requests per 5 min)")

        # Get current API key
        current_api_key = semantic_scholar.get_api_key()

        if current_api_key:
            st.success("✓ API key is set")
            st.caption("You have access to higher rate limits (5,000 requests per 5 minutes)")

            col_key1, col_key2 = st.columns([3, 1])
            with col_key1:
                # Show masked key
                masked_key = current_api_key[:8] + "*" * (len(current_api_key) - 12) + current_api_key[-4:] if len(current_api_key) > 12 else "*" * len(current_api_key)
                st.code(masked_key)
            with col_key2:
                if st.button("Remove Key", type="secondary", use_container_width=True):
                    semantic_scholar.set_api_key("")
                    st.success("API key removed")
                    st.rerun()
        else:
            st.info("⚠️ No API key set. Using unauthenticated access (100 requests per 5 minutes)")
            st.caption("Get a free API key at: https://www.semanticscholar.org/product/api")

            # Input for new key
            with st.form("semantic_scholar_api_key_form"):
                new_api_key = st.text_input(
                    "Enter your API key:",
                    type="password",
                    placeholder="Paste your Semantic Scholar API key here"
                )

                col_submit, col_info = st.columns([1, 2])
                with col_submit:
                    if st.form_submit_button("Save API Key", type="primary", use_container_width=True):
                        if new_api_key.strip():
                            success = semantic_scholar.set_api_key(new_api_key.strip())
                            if success:
                                st.success("✓ API key saved!")
                                st.rerun()
                            else:
                                st.error("Failed to save API key")
                        else:
                            st.warning("Please enter an API key")

                with col_info:
                    st.caption("Your key is stored locally in data/settings.json")

        st.divider()

        # Collections Management
        st.markdown("#### 📁 Collections Management")
        st.caption("Create, rename, and manage collections for organizing your papers.")

        # List all collections
        all_collections_settings = collections.get_all_collections()

        if all_collections_settings:
            st.markdown(f"**Current Collections ({len(all_collections_settings)}):**")

            for coll in all_collections_settings:
                with st.container():
                    col_display, col_edit, col_delete = st.columns([3, 1, 1])

                    with col_display:
                        color = coll.get('color', '#6c757d')
                        st.markdown(
                            f'<div style="display: flex; align-items: center; gap: 10px; padding: 8px 0;">'
                            f'<span style="display: inline-block; background-color: {color}; color: white; '
                            f'padding: 6px 14px; border-radius: 12px; font-size: 14px;">{coll["name"]}</span>'
                            f'<span style="color: #666; font-size: 13px;">({coll["paper_count"]} papers)</span>'
                            f'</div>',
                            unsafe_allow_html=True
                        )

                    with col_edit:
                        if st.button("✏️", key=f"edit_coll_{coll['id']}", help="Rename collection"):
                            st.session_state[f"editing_collection_{coll['id']}"] = True
                            st.rerun()

                    with col_delete:
                        if st.button("🗑️", key=f"delete_coll_{coll['id']}", help="Delete collection", type="secondary"):
                            # Delete confirmation
                            if st.session_state.get(f"confirm_delete_coll_{coll['id']}", False):
                                result = collections.delete_collection(coll['id'])
                                if result['success']:
                                    st.toast(result['message'], icon="✅")
                                    st.session_state[f"confirm_delete_coll_{coll['id']}"] = False
                                    st.rerun()
                                else:
                                    st.error(result['message'])
                            else:
                                st.session_state[f"confirm_delete_coll_{coll['id']}"] = True
                                st.rerun()

                    # Show edit form if editing
                    if st.session_state.get(f"editing_collection_{coll['id']}", False):
                        with st.form(key=f"edit_form_{coll['id']}"):
                            st.caption(f"**Editing: {coll['name']}**")
                            new_name = st.text_input("New Name", value=coll['name'], key=f"new_name_{coll['id']}")
                            new_color = st.color_picker("Color", value=coll.get('color', '#6c757d'), key=f"new_color_{coll['id']}")
                            new_desc = st.text_area("Description", value=coll.get('description', ''), key=f"new_desc_{coll['id']}")

                            col_save, col_cancel = st.columns(2)
                            with col_save:
                                if st.form_submit_button("💾 Save", use_container_width=True, type="primary"):
                                    result = collections.update_collection(
                                        coll['id'],
                                        name=new_name.strip() if new_name.strip() != coll['name'] else None,
                                        color=new_color if new_color != coll.get('color') else None,
                                        description=new_desc.strip() if new_desc.strip() != coll.get('description', '') else None
                                    )
                                    if result['success']:
                                        st.toast("Collection updated!", icon="✅")
                                        st.session_state[f"editing_collection_{coll['id']}"] = False
                                        st.rerun()
                                    else:
                                        st.error(result['message'])
                            with col_cancel:
                                if st.form_submit_button("❌ Cancel", use_container_width=True):
                                    st.session_state[f"editing_collection_{coll['id']}"] = False
                                    st.rerun()

                    # Show delete confirmation message
                    if st.session_state.get(f"confirm_delete_coll_{coll['id']}", False):
                        st.warning(f"⚠️ Click delete again to confirm removing '{coll['name']}' (papers will not be deleted)")

                    st.markdown("---")
        else:
            st.info("No collections yet. Create one below.")

        # Create new collection form
        with st.expander("➕ Create New Collection", expanded=False):
            with st.form(key="create_new_collection_form"):
                new_coll_name = st.text_input(
                    "Collection Name",
                    placeholder="e.g., SOH Methods, Grant Proposal, EIS Papers"
                )
                new_coll_color = st.color_picker("Color", value="#6c757d")
                new_coll_desc = st.text_area(
                    "Description (optional)",
                    placeholder="Brief description of this collection...",
                    height=80
                )

                if st.form_submit_button("Create Collection", type="primary", use_container_width=True):
                    if new_coll_name.strip():
                        result = collections.create_collection(
                            new_coll_name.strip(),
                            new_coll_color,
                            new_coll_desc.strip()
                        )
                        if result['success']:
                            st.toast(f"Collection '{new_coll_name}' created!", icon="✅")
                            st.rerun()
                        else:
                            st.error(result['message'])
                    else:
                        st.warning("Please enter a collection name")

        st.divider()

        # About section
        st.markdown("#### About")
        st.markdown("""
        **Astrolabe Paper Database**

        A comprehensive research paper management system with:
        - PDF ingestion and text extraction
        - Semantic search powered by ChromaDB
        - Metadata extraction and management
        - Reading history tracking
        - Automatic backups

        For more information, see the documentation files in the project directory.
        """)

    # ===== TAB 7: TRASH =====
    with tab7:
        st.session_state.active_tab = "Trash"

        st.markdown("### 🗑️ Trash")
        st.caption("Papers moved to trash. Restore them or permanently delete after review.")

        # Get trash papers
        trash_papers = get_trash_papers()

        if len(trash_papers) == 0:
            st.info("🎉 Trash is empty!")
        else:
            st.caption(f"**{len(trash_papers)} paper(s) in trash**")

            # Empty Trash button
            col1, col2 = st.columns([1, 3])
            with col1:
                if st.button("🗑️ Empty Trash", type="secondary", use_container_width=True):
                    if 'confirm_empty_trash' not in st.session_state:
                        st.session_state.confirm_empty_trash = False

                    if st.session_state.confirm_empty_trash:
                        # Permanently delete all trash papers
                        success_count = 0
                        for paper in trash_papers:
                            result = permanently_delete_paper(paper['filename'])
                            if result['success']:
                                success_count += 1

                        st.session_state.confirm_empty_trash = False
                        st.toast(f"🗑️ Permanently deleted {success_count} paper(s)", icon="✅")
                        st.rerun()
                    else:
                        st.session_state.confirm_empty_trash = True
                        st.warning("⚠️ Click again to confirm. This will permanently delete all trash papers.")

            with col2:
                st.caption("Permanently delete all papers in trash")

            st.divider()

            # Display trash papers
            for paper in trash_papers:
                with st.expander(f"📄 {paper['title']}", expanded=False):
                    # Paper details
                    col_info, col_actions = st.columns([3, 1])

                    with col_info:
                        if paper['authors']:
                            if isinstance(paper['authors'], list):
                                authors_str = '; '.join(paper['authors'])
                            else:
                                authors_str = paper['authors']
                            st.caption(f"**Authors:** {authors_str}")

                        if paper['year']:
                            st.caption(f"**Year:** {paper['year']}")

                        if paper['journal']:
                            st.caption(f"**Journal:** {paper['journal']}")

                        if paper['deleted_at']:
                            st.caption(f"**Deleted:** {paper['deleted_at']}")

                    with col_actions:
                        # Restore button
                        if st.button("♻️ Restore", key=f"restore_{paper['filename']}", use_container_width=True):
                            result = restore_paper(paper['filename'])
                            if result['success']:
                                st.toast(result['message'], icon="✅")
                                st.rerun()
                            else:
                                st.error(result['message'])

                        # Permanent delete button
                        if st.button("🗑️ Delete Forever", key=f"delete_forever_{paper['filename']}", type="secondary", use_container_width=True):
                            # Confirmation
                            confirm_key = f"confirm_delete_forever_{paper['filename']}"
                            if st.session_state.get(confirm_key, False):
                                result = permanently_delete_paper(paper['filename'])
                                if result['success']:
                                    st.toast(result['message'], icon="✅")
                                    st.session_state[confirm_key] = False
                                    st.rerun()
                                else:
                                    st.error(result['message'])
                            else:
                                st.session_state[confirm_key] = True
                                st.warning("⚠️ Click again to confirm permanent deletion")

            st.divider()

            # Auto-cleanup info
            st.caption("💡 **Tip:** Papers are automatically deleted permanently after 30 days in trash.")

            # Manual trigger for auto-cleanup
            if st.button("🧹 Run Auto-Cleanup Now (delete papers >30 days old)", use_container_width=False):
                result = auto_cleanup_old_trash(days=30)
                if result['deleted_count'] > 0:
                    st.toast(f"🧹 Cleaned up {result['deleted_count']} old paper(s)", icon="✅")
                    st.rerun()
                else:
                    st.info("No papers older than 30 days found in trash")


def import_reference(ref_data: dict) -> Dict[str, Any]:
    """
    Import a reference paper into the library.
    ALWAYS saves metadata, tries to get PDF if possible.

    Args:
        ref_data: Reference dictionary with DOI, title, authors, etc.

    Returns:
        Dict with success status and message
    """
    import json
    from pathlib import Path
    from datetime import datetime
    import requests
    import logging

    logger = logging.getLogger(__name__)
    result = {'success': False, 'message': '', 'filename': None}

    try:
        # Normalize DOI function (same as in references table)
        def normalize_doi(doi_string):
            if not doi_string:
                return ''
            doi = doi_string.lower().strip()
            for prefix in ['https://doi.org/', 'http://doi.org/', 'doi.org/', 'doi:']:
                if doi.startswith(prefix):
                    doi = doi[len(prefix):]
            return doi

        doi = normalize_doi(ref_data.get('DOI', ''))
        title = ref_data.get('article-title', 'Untitled')

        # Create filename
        if doi:
            filename = doi.replace('/', '_').replace('.', '_') + '.pdf'
        else:
            safe_title = ''.join(c for c in title[:50] if c.isalnum() or c in (' ', '-', '_'))
            filename = safe_title.replace(' ', '_') + '.pdf'

        # Load existing metadata
        metadata_file = Path("data/metadata.json")
        metadata_file.parent.mkdir(parents=True, exist_ok=True)

        if metadata_file.exists():
            with open(metadata_file, 'r', encoding='utf-8') as f:
                all_metadata = json.load(f)
        else:
            all_metadata = {}

        # Check if already exists by DOI (consistent with references table)
        if doi:
            # Check if any existing paper has this DOI
            for existing_filename, existing_meta in all_metadata.items():
                existing_doi = normalize_doi(existing_meta.get('doi', ''))
                if existing_doi and existing_doi == doi:
                    result['success'] = True
                    result['message'] = f"Already in library (matched DOI: {doi})"
                    result['debug_matched_file'] = existing_filename
                    return result

        # Also check by filename as fallback
        if filename in all_metadata:
            result['success'] = True
            result['message'] = f"Already in library (filename: {filename})"
            result['debug_filename'] = filename
            return result

        # Start with reference data
        metadata = {
            'title': title,
            'authors': [ref_data.get('author', '')] if ref_data.get('author') else [],
            'year': ref_data.get('year', ''),
            'journal': ref_data.get('journal-title', ''),
            'doi': doi,
            'volume': ref_data.get('volume', ''),
            'issue': ref_data.get('issue', ''),
            'pages': ref_data.get('first-page', ''),
            'date_added': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'extracted_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'chemistries': [],
            'topics': [],
            'application': 'general',
            'paper_type': 'reference',
            'author_keywords': [],
            'abstract': '',
            'references': [],
            'pdf_status': 'needs_pdf'  # Default: no PDF
        }

        # If has DOI, enrich with CrossRef data
        if doi:
            crossref_data = query_crossref_for_metadata(doi)
            if crossref_data:
                # Update with richer CrossRef data
                metadata.update({
                    'title': crossref_data.get('title', title),
                    'authors': crossref_data.get('authors', metadata['authors']),
                    'year': crossref_data.get('year', metadata['year']),
                    'journal': crossref_data.get('journal', metadata['journal']),
                    'volume': crossref_data.get('volume', metadata['volume']),
                    'issue': crossref_data.get('issue', metadata['issue']),
                    'pages': crossref_data.get('pages', metadata['pages']),
                    'abstract': crossref_data.get('abstract', ''),
                    'author_keywords': crossref_data.get('author_keywords', []),
                    'references': crossref_data.get('references', [])
                })

            # Try Unpaywall for PDF
            try:
                unpaywall_url = f"https://api.unpaywall.org/v2/{doi}?email=researcher@example.com"
                response = requests.get(unpaywall_url, timeout=10)
                if response.status_code == 200:
                    data = response.json()
                    pdf_url = data.get('best_oa_location', {}).get('url_for_pdf')

                    if pdf_url:
                        # Download PDF
                        pdf_response = requests.get(pdf_url, timeout=30)
                        if pdf_response.status_code == 200:
                            papers_dir = Path("papers")
                            papers_dir.mkdir(parents=True, exist_ok=True)
                            pdf_path = papers_dir / filename

                            with open(pdf_path, 'wb') as f:
                                f.write(pdf_response.content)

                            metadata['pdf_status'] = 'has_pdf'
                            result['message'] = "Added with PDF"
                        else:
                            result['message'] = "Added (metadata only, PDF unavailable)"
                    else:
                        result['message'] = "Added (metadata only, no open access PDF)"
                else:
                    result['message'] = "Added (metadata only)"
            except:
                result['message'] = "Added (metadata only, PDF check failed)"
        else:
            result['message'] = "Added (metadata only, no DOI)"

        # Save metadata (ALWAYS, regardless of PDF status)
        all_metadata[filename] = metadata
        with open(metadata_file, 'w', encoding='utf-8') as f:
            json.dump(all_metadata, f, indent=2, ensure_ascii=False)

        result['success'] = True
        result['filename'] = filename

    except Exception as e:
        result['message'] = f"Error: {str(e)}"

    return result


def process_url_import(url: str, progress_container) -> Dict[str, Any]:
    """
    Import a paper from URL (arXiv, DOI, or direct PDF link).

    Args:
        url: URL to import from
        progress_container: Streamlit container for progress updates

    Returns:
        Dictionary with import results
    """
    import subprocess
    import urllib.parse
    from pathlib import Path

    papers_dir = Path("papers")
    papers_dir.mkdir(parents=True, exist_ok=True)

    result = {
        'success': False,
        'title': None,
        'filename': None,
        'error': None,
        'metadata_only': False
    }

    url = url.strip()

    with progress_container:
        st.info(f"🔗 Processing URL: {url}")

        try:
            # Detect URL type
            if 'arxiv.org' in url:
                # arXiv link
                st.caption("📄 Detected: arXiv paper")

                # Extract arXiv ID
                arxiv_match = re.search(r'arxiv\.org/(?:abs|pdf)/(\d+\.\d+)', url)
                if not arxiv_match:
                    result['error'] = "Invalid arXiv URL format"
                    return result

                arxiv_id = arxiv_match.group(1)
                pdf_url = f"https://arxiv.org/pdf/{arxiv_id}.pdf"
                filename = f"arxiv_{arxiv_id.replace('.', '_')}.pdf"

                st.caption(f"📥 Downloading from arXiv (ID: {arxiv_id})...")

                # Download PDF
                response = requests.get(pdf_url, timeout=30)
                if response.status_code == 200:
                    filepath = papers_dir / filename
                    with open(filepath, 'wb') as f:
                        f.write(response.content)

                    result['filename'] = filename
                    result['success'] = True
                    st.caption(f"✓ Downloaded: {filename}")
                else:
                    result['error'] = f"Failed to download from arXiv (HTTP {response.status_code})"
                    return result

            elif any(publisher in url.lower() for publisher in [
                'sciencedirect.com', 'ieeexplore.ieee.org', 'onlinelibrary.wiley.com',
                'link.springer.com', 'nature.com/articles', 'mdpi.com', 'cell.com',
                'thelancet.com', 'pubs.acs.org', 'pubs.rsc.org', 'iopscience.iop.org'
            ]):
                # Publisher article page
                st.caption("📰 Detected: Publisher article page")
                st.caption(f"🔍 Extracting DOI from page...")

                doi = None

                # Try to extract DOI from URL pattern first
                if 'doi.org' in url or '/doi/' in url:
                    # DOI is in the URL
                    doi_match = re.search(r'(?:doi\.org/|/doi/(?:abs/|full/)?)(10\.\d+/[^\s?&#]+)', url)
                    if doi_match:
                        doi = doi_match.group(1)

                # If not in URL, scrape from page
                if not doi:
                    try:
                        st.caption("🌐 Fetching page to extract DOI...")
                        # Use more complete browser headers to avoid blocking
                        headers = {
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                            'Accept-Language': 'en-US,en;q=0.9',
                            'Accept-Encoding': 'gzip, deflate, br',
                            'DNT': '1',
                            'Connection': 'keep-alive',
                            'Upgrade-Insecure-Requests': '1',
                            'Sec-Fetch-Dest': 'document',
                            'Sec-Fetch-Mode': 'navigate',
                            'Sec-Fetch-Site': 'none',
                            'Cache-Control': 'max-age=0'
                        }

                        page_response = requests.get(url, timeout=15, headers=headers)

                        if page_response.status_code == 403:
                            st.warning("⚠️ Publisher blocked automated access (403 Forbidden)")
                            st.info("💡 Workaround: Manually enter the DOI instead, or download the PDF and upload it.")
                            result['error'] = "Publisher blocked automated access. Try entering DOI directly or upload PDF."
                            return result
                        elif page_response.status_code != 200:
                            st.warning(f"⚠️ Could not fetch page (HTTP {page_response.status_code})")
                            result['error'] = f"HTTP {page_response.status_code} when fetching page"
                            return result

                        if page_response.status_code == 200:
                            from bs4 import BeautifulSoup
                            soup = BeautifulSoup(page_response.text, 'html.parser')

                            # Try all common meta tag patterns
                            meta_tags_to_try = [
                                ('name', 'citation_doi'),
                                ('name', 'DC.Identifier'),
                                ('property', 'citation_doi'),
                                ('name', 'DOI'),
                                ('name', 'dc.identifier'),
                                ('property', 'og:identifier'),
                                ('name', 'prism.doi'),  # Common in ScienceDirect
                            ]

                            for attr, value in meta_tags_to_try:
                                doi_meta = soup.find('meta', {attr: value})
                                if doi_meta and doi_meta.get('content'):
                                    doi_content = doi_meta['content'].strip()
                                    # Extract just the DOI part
                                    if 'doi.org/' in doi_content:
                                        doi = doi_content.split('doi.org/')[-1]
                                    elif doi_content.startswith('10.'):
                                        doi = doi_content

                                    if doi:
                                        st.caption(f"✓ Found DOI in meta tag: {attr}={value}")
                                        break

                            # If still no DOI, search page HTML for DOI pattern
                            if not doi:
                                # Look for DOI in script tags (ScienceDirect often has it in JSON-LD)
                                script_tags = soup.find_all('script', {'type': 'application/ld+json'})
                                for script in script_tags:
                                    doi_match = re.search(r'"doi"\s*:\s*"(10\.\d+/[^"]+)"', script.string or '')
                                    if doi_match:
                                        doi = doi_match.group(1)
                                        st.caption("✓ Found DOI in JSON-LD schema")
                                        break

                            # Last resort: search entire page text for DOI pattern
                            if not doi:
                                doi_pattern = re.search(r'\b(10\.\d{4,}/[^\s<>"\']+)\b', page_response.text)
                                if doi_pattern:
                                    candidate = doi_pattern.group(1)
                                    # Clean up common trailing characters
                                    candidate = re.sub(r'[,;.\)]+$', '', candidate)
                                    if candidate:
                                        doi = candidate
                                        st.caption("✓ Found DOI in page content")

                    except Exception as e:
                        st.warning(f"⚠️ Could not fetch page: {str(e)}")

                if not doi:
                    result['error'] = "Could not extract DOI from publisher page"
                    return result

                st.caption(f"✓ Found DOI: {doi}")

                # Now proceed with DOI-based lookup
                st.caption(f"📖 Looking up metadata for DOI: {doi}")

                # Get metadata from CrossRef
                metadata = query_crossref_for_metadata(doi)
                if not metadata:
                    result['error'] = "Could not retrieve metadata from CrossRef"
                    return result

                result['title'] = metadata.get('title', 'Unknown')
                st.caption(f"✓ Found: {result['title']}")

                # Try to find open access PDF via Unpaywall
                st.caption("🔓 Checking for open access PDF via Unpaywall...")

                unpaywall_url = f"https://api.unpaywall.org/v2/{doi}?email=researcher@example.com"
                unpaywall_response = requests.get(unpaywall_url, timeout=10)

                pdf_url = None
                if unpaywall_response.status_code == 200:
                    unpaywall_data = unpaywall_response.json()
                    if unpaywall_data.get('is_oa') and unpaywall_data.get('best_oa_location'):
                        pdf_url = unpaywall_data['best_oa_location'].get('url_for_pdf')

                if pdf_url:
                    st.caption(f"✓ Found open access PDF!")
                    st.caption(f"📥 Downloading from {urllib.parse.urlparse(pdf_url).netloc}...")

                    # Download PDF
                    try:
                        pdf_response = requests.get(pdf_url, timeout=30, allow_redirects=True)
                        if pdf_response.status_code == 200 and pdf_response.headers.get('content-type', '').startswith('application/pdf'):
                            # Create safe filename from DOI
                            safe_doi = doi.replace('/', '_').replace('.', '_')
                            filename = f"doi_{safe_doi}.pdf"
                            filepath = papers_dir / filename

                            with open(filepath, 'wb') as f:
                                f.write(pdf_response.content)

                            result['filename'] = filename
                            result['success'] = True
                            st.caption(f"✓ Downloaded: {filename}")
                        else:
                            st.warning("⚠️ Could not download PDF (may be paywalled)")
                            result['metadata_only'] = True
                            result['filename'] = save_metadata_only_paper(doi, metadata)
                            result['success'] = True
                    except Exception as e:
                        st.warning(f"⚠️ PDF download failed: {str(e)}")
                        result['metadata_only'] = True
                        result['filename'] = save_metadata_only_paper(doi, metadata)
                        result['success'] = True
                else:
                    st.warning("⚠️ No open access PDF found - this paper may be paywalled")
                    result['metadata_only'] = True
                    result['filename'] = save_metadata_only_paper(doi, metadata)
                    result['success'] = True

            elif 'doi.org' in url or url.startswith('10.'):
                # DOI link or DOI string
                st.caption("🔍 Detected: DOI")

                # Extract DOI
                if url.startswith('10.'):
                    doi = url
                else:
                    doi = url.split('doi.org/')[-1]

                st.caption(f"📖 Looking up metadata for DOI: {doi}")

                # Get metadata from CrossRef
                metadata = query_crossref_for_metadata(doi)
                if not metadata:
                    result['error'] = "Could not retrieve metadata from CrossRef"
                    return result

                result['title'] = metadata.get('title', 'Unknown')
                st.caption(f"✓ Found: {result['title']}")

                # Try to find open access PDF via Unpaywall
                st.caption("🔓 Checking for open access PDF via Unpaywall...")

                unpaywall_url = f"https://api.unpaywall.org/v2/{doi}?email=researcher@example.com"
                unpaywall_response = requests.get(unpaywall_url, timeout=10)

                pdf_url = None
                if unpaywall_response.status_code == 200:
                    unpaywall_data = unpaywall_response.json()
                    if unpaywall_data.get('is_oa') and unpaywall_data.get('best_oa_location'):
                        pdf_url = unpaywall_data['best_oa_location'].get('url_for_pdf')

                if pdf_url:
                    st.caption(f"✓ Found open access PDF!")
                    st.caption(f"📥 Downloading from {urllib.parse.urlparse(pdf_url).netloc}...")

                    # Download PDF
                    try:
                        pdf_response = requests.get(pdf_url, timeout=30, allow_redirects=True)
                        if pdf_response.status_code == 200 and pdf_response.headers.get('content-type', '').startswith('application/pdf'):
                            # Create safe filename from DOI
                            safe_doi = doi.replace('/', '_').replace('.', '_')
                            filename = f"doi_{safe_doi}.pdf"
                            filepath = papers_dir / filename

                            with open(filepath, 'wb') as f:
                                f.write(pdf_response.content)

                            result['filename'] = filename
                            result['success'] = True
                            st.caption(f"✓ Downloaded: {filename}")
                        else:
                            st.warning("⚠️ Could not download PDF (may be paywalled)")
                            result['metadata_only'] = True
                            result['filename'] = save_metadata_only_paper(doi, metadata)
                            result['success'] = True
                    except Exception as e:
                        st.warning(f"⚠️ PDF download failed: {str(e)}")
                        result['metadata_only'] = True
                        result['filename'] = save_metadata_only_paper(doi, metadata)
                        result['success'] = True
                else:
                    st.warning("⚠️ No open access PDF found - this paper may be paywalled")
                    result['metadata_only'] = True
                    result['filename'] = save_metadata_only_paper(doi, metadata)
                    result['success'] = True

            elif url.endswith('.pdf') or 'pdf' in url.lower():
                # Direct PDF link
                st.caption("📄 Detected: Direct PDF link")
                st.caption(f"📥 Downloading PDF...")

                # Download PDF
                response = requests.get(url, timeout=30, allow_redirects=True)
                if response.status_code == 200:
                    # Try to get filename from URL or Content-Disposition header
                    filename = None
                    if 'content-disposition' in response.headers:
                        cd = response.headers['content-disposition']
                        filename_match = re.findall('filename="?([^"]+)"?', cd)
                        if filename_match:
                            filename = filename_match[0]

                    if not filename:
                        # Extract from URL
                        filename = url.split('/')[-1].split('?')[0]
                        if not filename.endswith('.pdf'):
                            filename = f"downloaded_{int(time.time())}.pdf"

                    filepath = papers_dir / filename
                    with open(filepath, 'wb') as f:
                        f.write(response.content)

                    result['filename'] = filename
                    result['success'] = True
                    st.caption(f"✓ Downloaded: {filename}")
                else:
                    result['error'] = f"Failed to download PDF (HTTP {response.status_code})"
                    return result
            else:
                result['error'] = "Unrecognized URL format. Supported: arXiv, DOI (doi.org/...), or direct PDF links"
                return result

            # Run ingestion pipeline if we have a PDF
            if result['filename'] and not result['metadata_only']:
                st.divider()
                st.info(f"📊 Processing paper through pipeline...")

                progress_bar = st.progress(0)
                status_text = st.empty()

                try:
                    # Stage 1: Parse
                    status_text.text("Stage 1/4: Extracting text from PDF...")
                    progress_bar.progress(25)
                    subprocess.run(
                        [sys.executable, "scripts/ingest_pipeline.py", "--stage", "parse", "--new-only"],
                        check=True,
                        capture_output=True,
                        text=True
                    )

                    # Stage 2: Chunk
                    status_text.text("Stage 2/4: Creating chunks...")
                    progress_bar.progress(50)
                    subprocess.run(
                        [sys.executable, "scripts/ingest_pipeline.py", "--stage", "chunk", "--new-only"],
                        check=True,
                        capture_output=True,
                        text=True
                    )

                    # Stage 3: Metadata
                    status_text.text("Stage 3/4: Extracting metadata...")
                    progress_bar.progress(75)
                    subprocess.run(
                        [sys.executable, "scripts/ingest_pipeline.py", "--stage", "metadata", "--new-only"],
                        check=True,
                        capture_output=True,
                        text=True
                    )

                    # Stage 4: Embed
                    status_text.text("Stage 4/4: Creating embeddings...")
                    progress_bar.progress(90)
                    subprocess.run(
                        [sys.executable, "scripts/ingest_pipeline.py", "--stage", "embed"],
                        check=True,
                        capture_output=True,
                        text=True
                    )

                    progress_bar.progress(100)
                    status_text.text("✅ Processing complete!")

                except subprocess.CalledProcessError as e:
                    result['error'] = f"Pipeline processing failed: {str(e)}"
                    return result

        except requests.exceptions.Timeout:
            result['error'] = "Request timed out - server not responding"
        except requests.exceptions.ConnectionError:
            result['error'] = "Connection error - check your internet connection"
        except Exception as e:
            result['error'] = f"Unexpected error: {str(e)}"

    return result


def process_uploaded_pdfs(uploaded_files: list, progress_container, replace_duplicates: bool = False) -> Dict[str, Any]:
    """
    Process uploaded PDF files through the ingestion pipeline.

    Args:
        uploaded_files: List of uploaded file objects from st.file_uploader
        progress_container: Streamlit container for progress updates
        replace_duplicates: If True, replace existing files; if False, skip them

    Returns:
        Dictionary with processing results
    """
    import subprocess
    from pathlib import Path

    papers_dir = Path("papers")
    papers_dir.mkdir(parents=True, exist_ok=True)

    results = {
        'saved': [],
        'replaced': [],
        'skipped': [],
        'failed': [],
        'total': len(uploaded_files)
    }

    with progress_container:
        # Show immediate feedback
        st.info(f"🚀 Starting to process {len(uploaded_files)} file(s)...")
        time.sleep(0.5)  # Brief pause so user sees the message

    # Save uploaded files
    for i, uploaded_file in enumerate(uploaded_files, 1):
        filename = uploaded_file.name
        target_path = papers_dir / filename
        is_replacement = False

        with progress_container:
            st.caption(f"📄 Saving file {i}/{len(uploaded_files)}: {filename}")

        # Check for duplicates
        if target_path.exists():
            if replace_duplicates:
                # Delete the old file and mark for replacement
                try:
                    target_path.unlink()
                    is_replacement = True
                    with progress_container:
                        st.caption(f"♻️ Replacing existing file: {filename}")
                except Exception as e:
                    results['failed'].append((filename, f"Failed to replace: {str(e)}"))
                    with progress_container:
                        st.error(f"❌ Failed to replace {filename}: {str(e)}")
                    continue
            else:
                results['skipped'].append(filename)
                with progress_container:
                    st.caption(f"⏭️ Skipping duplicate: {filename}")
                continue

        try:
            # Save the file
            with open(target_path, 'wb') as f:
                f.write(uploaded_file.getbuffer())

            if is_replacement:
                results['replaced'].append(filename)
            else:
                results['saved'].append(filename)

            with progress_container:
                st.caption(f"✓ Saved: {filename}")

        except Exception as e:
            results['failed'].append((filename, str(e)))
            with progress_container:
                st.error(f"❌ Failed to save {filename}: {str(e)}")

    # Run ingestion pipeline if we saved or replaced any files
    total_to_process = len(results['saved']) + len(results['replaced'])
    if total_to_process > 0:
        all_papers = results['saved'] + results['replaced']

        with progress_container:
            st.divider()
            if results['replaced']:
                st.info(f"📊 Processing {total_to_process} paper(s) through pipeline ({len(results['replaced'])} replacement(s))...")
            else:
                st.info(f"📊 Processing {total_to_process} new paper(s) through pipeline...")

            overall_progress = st.progress(0)
            stage_status = st.empty()
            paper_status = st.empty()

            try:
                # Stage 1: Parse (Extract text)
                stage_status.markdown("**Stage 1/4: 📄 Extracting text from PDFs**")
                for i, paper in enumerate(all_papers, 1):
                    paper_status.text(f"   Processing paper {i}/{total_to_process}: {paper}")
                    overall_progress.progress(int((i / total_to_process) * 20))
                    time.sleep(0.1)  # Brief pause for visibility

                result = subprocess.run(
                    [sys.executable, "scripts/ingest_pipeline.py", "--stage", "parse", "--new-only"],
                    capture_output=True,
                    text=True
                )
                if result.returncode != 0:
                    raise subprocess.CalledProcessError(result.returncode, result.args, result.stdout, result.stderr)

                overall_progress.progress(25)
                paper_status.text("   ✓ Text extraction complete")
                time.sleep(0.3)

                # Stage 2: Chunk
                stage_status.markdown("**Stage 2/4: 📑 Creating chunks**")
                for i, paper in enumerate(all_papers, 1):
                    paper_status.text(f"   Chunking paper {i}/{total_to_process}: {paper}")
                    overall_progress.progress(25 + int((i / total_to_process) * 20))
                    time.sleep(0.1)

                result = subprocess.run(
                    [sys.executable, "scripts/ingest_pipeline.py", "--stage", "chunk", "--new-only"],
                    capture_output=True,
                    text=True
                )
                if result.returncode != 0:
                    raise subprocess.CalledProcessError(result.returncode, result.args, result.stdout, result.stderr)

                overall_progress.progress(50)
                paper_status.text("   ✓ Chunking complete")
                time.sleep(0.3)

                # Stage 3: Metadata
                stage_status.markdown("**Stage 3/4: 🔍 Extracting metadata**")
                paper_status.text("   Using Claude to analyze papers and extract metadata...")
                overall_progress.progress(55)

                for i, paper in enumerate(all_papers, 1):
                    paper_status.text(f"   Analyzing paper {i}/{total_to_process}: {paper}")
                    # This stage takes longer, so update less frequently
                    overall_progress.progress(55 + int((i / total_to_process) * 20))
                    time.sleep(0.2)

                result = subprocess.run(
                    [sys.executable, "scripts/ingest_pipeline.py", "--stage", "metadata", "--new-only"],
                    capture_output=True,
                    text=True
                )
                if result.returncode != 0:
                    raise subprocess.CalledProcessError(result.returncode, result.args, result.stdout, result.stderr)

                overall_progress.progress(75)
                paper_status.text("   ✓ Metadata extraction complete")
                time.sleep(0.3)

                # Stage 4: Embed
                stage_status.markdown("**Stage 4/4: 🧮 Generating embeddings and indexing**")
                paper_status.text("   Creating vector embeddings for semantic search...")
                overall_progress.progress(80)

                for i, paper in enumerate(all_papers, 1):
                    paper_status.text(f"   Embedding paper {i}/{total_to_process}: {paper}")
                    overall_progress.progress(80 + int((i / total_to_process) * 15))
                    time.sleep(0.1)

                result = subprocess.run(
                    [sys.executable, "scripts/ingest_pipeline.py", "--stage", "embed"],
                    capture_output=True,
                    text=True
                )
                if result.returncode != 0:
                    raise subprocess.CalledProcessError(result.returncode, result.args, result.stdout, result.stderr)

                overall_progress.progress(100)
                paper_status.text("   ✓ Embeddings generated and indexed")
                stage_status.markdown("**✅ All stages complete!**")
                time.sleep(0.5)

            except subprocess.CalledProcessError as e:
                stage_status.markdown("**❌ Pipeline Error**")
                paper_status.text("")
                error_msg = e.stderr if e.stderr else str(e)
                st.error(f"Pipeline failed: {error_msg}")

                # Show detailed error if available
                if e.stdout:
                    with st.expander("Show pipeline output"):
                        st.code(e.stdout)

                # Mark all as failed
                all_files = results['saved'] + results['replaced']
                for filename in all_files:
                    results['failed'].append((filename, "Pipeline processing failed"))
                results['saved'] = []
                results['replaced'] = []

    return results


def display_query_card(query: Dict):
    """Display a single query card in the history view."""
    # Parse timestamp
    try:
        dt = datetime.fromisoformat(query['timestamp'])
        time_str = dt.strftime("%B %d, %Y at %I:%M %p")
    except:
        time_str = query['timestamp']

    # Create card layout
    col1, col2, col3, col4 = st.columns([6, 1, 1, 1])

    with col1:
        # Question preview (truncate if too long)
        question_preview = query['question']
        if len(question_preview) > 100:
            question_preview = question_preview[:100] + "..."

        st.markdown(f"**Q:** {question_preview}")
        st.caption(f"🕐 {time_str}")

        # Show filters if any
        if query['filters'] and any(query['filters'].values()):
            filter_text = ", ".join([
                f"{k.title()}: {v}"
                for k, v in query['filters'].items()
                if v
            ])
            st.caption(f"🔍 Filters: {filter_text}")

    with col2:
        # Star button
        star_icon = "⭐" if query['is_starred'] else "☆"
        if st.button(star_icon, key=f"star_{query['id']}", help="Star this query"):
            query_history.toggle_star(query['id'])
            st.rerun()

    with col3:
        # View button
        if st.button("👁️", key=f"view_{query['id']}", help="View this query"):
            # Load query into session state and switch to Research tab
            st.session_state.query_result = {
                'question': query['question'],
                'answer': query['answer'],
                'chunks': query['chunks'],
                'filters': query['filters'],
                'from_history': True
            }
            st.session_state.active_tab = "Research"
            st.rerun()

    with col4:
        # Delete button
        if st.button("🗑️", key=f"delete_{query['id']}", help="Delete this query"):
            query_history.delete_query(query['id'])
            st.toast("Query deleted", icon="🗑️")
            st.rerun()


# ============================================================================
# TRASH MANAGEMENT FUNCTIONS
# ============================================================================

def soft_delete_paper(filename: str) -> Dict[str, Any]:
    """
    Soft delete a paper by marking it as deleted and moving PDF to trash.

    Args:
        filename: Paper filename

    Returns:
        Dict with success status and message
    """
    import shutil
    from datetime import datetime

    try:
        # Load metadata
        metadata_file = Path("data/metadata.json")
        if not metadata_file.exists():
            return {'success': False, 'message': 'Metadata file not found'}

        with open(metadata_file, 'r', encoding='utf-8') as f:
            all_metadata = json.load(f)

        if filename not in all_metadata:
            return {'success': False, 'message': 'Paper not found in metadata'}

        # Mark as deleted in metadata
        all_metadata[filename]['deleted_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Save updated metadata
        with open(metadata_file, 'w', encoding='utf-8') as f:
            json.dump(all_metadata, f, indent=2, ensure_ascii=False)

        # Move PDF to trash folder if it exists
        pdf_path = Path("papers") / filename
        if pdf_path.exists():
            trash_dir = Path("papers/trash")
            trash_dir.mkdir(parents=True, exist_ok=True)

            trash_path = trash_dir / filename
            shutil.move(str(pdf_path), str(trash_path))

        # Remove from ChromaDB
        try:
            db = rag.DatabaseClient()
            db.collection.delete(where={"filename": filename})
        except Exception as e:
            # Non-fatal if ChromaDB deletion fails
            pass

        return {
            'success': True,
            'message': f'Moved "{all_metadata[filename].get("title", filename)}" to trash'
        }

    except Exception as e:
        return {'success': False, 'message': f'Error: {str(e)}'}


def restore_paper(filename: str) -> Dict[str, Any]:
    """
    Restore a paper from trash.

    Args:
        filename: Paper filename

    Returns:
        Dict with success status and message
    """
    import shutil

    try:
        # Load metadata
        metadata_file = Path("data/metadata.json")
        if not metadata_file.exists():
            return {'success': False, 'message': 'Metadata file not found'}

        with open(metadata_file, 'r', encoding='utf-8') as f:
            all_metadata = json.load(f)

        if filename not in all_metadata:
            return {'success': False, 'message': 'Paper not found in metadata'}

        # Remove deleted_at timestamp
        if 'deleted_at' in all_metadata[filename]:
            del all_metadata[filename]['deleted_at']

        # Save updated metadata
        with open(metadata_file, 'w', encoding='utf-8') as f:
            json.dump(all_metadata, f, indent=2, ensure_ascii=False)

        # Move PDF back from trash if it exists
        trash_path = Path("papers/trash") / filename
        if trash_path.exists():
            pdf_path = Path("papers") / filename
            pdf_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(trash_path), str(pdf_path))

        # Note: PDF will be re-indexed on next ingestion run
        # Or we could trigger re-ingestion here if needed

        return {
            'success': True,
            'message': f'Restored "{all_metadata[filename].get("title", filename)}" from trash'
        }

    except Exception as e:
        return {'success': False, 'message': f'Error: {str(e)}'}


def permanently_delete_paper(filename: str) -> Dict[str, Any]:
    """
    Permanently delete a paper (remove metadata, PDF, and ChromaDB entries).

    Args:
        filename: Paper filename

    Returns:
        Dict with success status and message
    """
    try:
        # Load metadata
        metadata_file = Path("data/metadata.json")
        if not metadata_file.exists():
            return {'success': False, 'message': 'Metadata file not found'}

        with open(metadata_file, 'r', encoding='utf-8') as f:
            all_metadata = json.load(f)

        if filename not in all_metadata:
            return {'success': False, 'message': 'Paper not found in metadata'}

        title = all_metadata[filename].get('title', filename)

        # Remove from metadata
        del all_metadata[filename]

        # Save updated metadata
        with open(metadata_file, 'w', encoding='utf-8') as f:
            json.dump(all_metadata, f, indent=2, ensure_ascii=False)

        # Delete PDF from trash
        trash_path = Path("papers/trash") / filename
        if trash_path.exists():
            trash_path.unlink()

        # Delete PDF from papers (if somehow still there)
        pdf_path = Path("papers") / filename
        if pdf_path.exists():
            pdf_path.unlink()

        # Remove from ChromaDB
        try:
            db = rag.DatabaseClient()
            db.collection.delete(where={"filename": filename})
        except Exception as e:
            # Non-fatal if ChromaDB deletion fails
            pass

        # Delete notes file if exists
        notes_file = Path(f"data/notes/{filename}.txt")
        if notes_file.exists():
            notes_file.unlink()

        return {
            'success': True,
            'message': f'Permanently deleted "{title}"'
        }

    except Exception as e:
        return {'success': False, 'message': f'Error: {str(e)}'}


def get_trash_papers() -> list:
    """
    Get list of papers in trash.

    Returns:
        List of paper metadata dicts for trashed papers
    """
    metadata_file = Path("data/metadata.json")
    if not metadata_file.exists():
        return []

    with open(metadata_file, 'r', encoding='utf-8') as f:
        all_metadata = json.load(f)

    # Filter for deleted papers
    trash_papers = []
    for filename, meta in all_metadata.items():
        if meta.get('deleted_at'):
            trash_papers.append({
                'filename': filename,
                'title': meta.get('title', filename),
                'authors': meta.get('authors', []),
                'year': meta.get('year', ''),
                'journal': meta.get('journal', ''),
                'deleted_at': meta.get('deleted_at', '')
            })

    # Sort by deleted_at (most recent first)
    trash_papers.sort(key=lambda x: x['deleted_at'], reverse=True)

    return trash_papers


def auto_cleanup_old_trash(days: int = 30) -> Dict[str, Any]:
    """
    Automatically delete papers that have been in trash for more than specified days.

    Args:
        days: Number of days after which to permanently delete

    Returns:
        Dict with count of deleted papers
    """
    from datetime import datetime, timedelta

    trash_papers = get_trash_papers()
    cutoff_date = datetime.now() - timedelta(days=days)

    deleted_count = 0
    for paper in trash_papers:
        deleted_at = datetime.strptime(paper['deleted_at'], "%Y-%m-%d %H:%M:%S")
        if deleted_at < cutoff_date:
            result = permanently_delete_paper(paper['filename'])
            if result['success']:
                deleted_count += 1

    return {
        'success': True,
        'deleted_count': deleted_count,
        'message': f'Auto-cleaned {deleted_count} papers older than {days} days'
    }


if __name__ == "__main__":
    main()
